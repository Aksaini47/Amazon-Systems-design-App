"""
Read bulk_upload_update_final_FIXED_v4.xlsx and produce v5 with TWO sheets:
  Sheet 1: 'Template'  — all 1283 SKUs (same as v4)
  Sheet 2: 'Flagged'   — only SKUs with detected issues, with all 28 original columns
                        plus 2 extra columns: 'issue_types' and 'issue_details'

Issue types detected:
  - no_model_codes              : BP1 has only the brand+model, no model code tokens
  - potential_device_mismatch   : title's numeric/key identifier missing from MongoDB matched device name
  - multi_variant_title         : title contains '/' (multi-device variants — lookup typically fails or picks wrong one)
  - bare_gold_in_keywords       : keywords field has bare 'gold' without parens
"""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from pymongo import MongoClient

SOURCE = 'bulk_upload_update_final_FIXED_v4.xlsx'
DEST   = 'bulk_upload_update_final_FIXED_v5.xlsx'
MONGO_URI = 'mongodb://dev_user:jfndACzw0ypeaNPi@ac-pn1ls1y-shard-00-00.wpdrnxc.mongodb.net:27017,ac-pn1ls1y-shard-00-01.wpdrnxc.mongodb.net:27017,ac-pn1ls1y-shard-00-02.wpdrnxc.mongodb.net:27017/repairfully?ssl=true&replicaSet=atlas-mwdddo-shard-0&authSource=admin'

# ---------- Load source ----------
print(f'Loading source: {SOURCE}')
src_wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
src_ws = src_wb['Template']
all_rows = list(src_ws.iter_rows(values_only=True))
header = list(all_rows[0])
data_rows = all_rows[1:]
print(f'Loaded {len(data_rows)} SKUs.')

# ---------- Connect to Mongo for accuracy check ----------
print('Loading MongoDB device lookup for accuracy check...')
db = MongoClient(MONGO_URI, serverSelectionTimeoutMS=15000)['repairfully']
coll = db['device_model_lookup']

# Pre-fetch only what we need
mongo_lookup = {}  # brandSlug -> list[(deviceNameLower, dn_norm, codes)]
for d in coll.find({}, {'brandSlug':1, 'deviceName':1, 'deviceNameLower':1, 'codes':1, 'finalAssignedCode':1}):
    slug = d.get('brandSlug') or ''
    dn = d.get('deviceName') or ''
    dn_lower = d.get('deviceNameLower') or ''
    codes = d.get('codes') or []
    if not codes and d.get('finalAssignedCode'):
        codes = [d['finalAssignedCode']]
    if slug and dn_lower:
        mongo_lookup.setdefault(slug, []).append((dn_lower, dn, codes))
print(f'Loaded {sum(len(v) for v in mongo_lookup.values())} devices')

BRAND_ALIASES = {'redmi': 'xiaomi', 'mi': 'xiaomi', 'iphone': 'apple', '1+': 'oneplus'}

def normalize(s):
    s = (s or '').lower()
    s = re.sub(r'[^a-z0-9 ]+', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

TITLE_HEAD_RE = re.compile(
    r'compatible\s+for\s+(.+?)\s+'
    r'(?:\(|with\s+frame|without\s+frame|fingerprint|lcd|oled|amoled|in[\s-]?cell|careog|super\s+(?:retina|oled|amoled)|display)',
    re.IGNORECASE
)
def parse_title(title):
    m = TITLE_HEAD_RE.search(title or '')
    head = m.group(1).strip() if m else (title or '').replace('Compatible for ', '').strip()
    head = re.sub(r'\s*\([^)]*\)\s*$', '', head).strip()
    parts = head.split()
    brand = parts[0] if parts else ''
    return brand, head

def extract_id_tokens(model_phrase):
    """Extract tokens from model phrase that contain digits OR are key words like 'pro', 'plus', 'max', 'play', 'note'.
    These are the identifiers that MUST appear in the matched device name."""
    norm = normalize(model_phrase)
    # First strip brand prefix
    for p in ['samsung','apple','motorola','oneplus','oppo','vivo','realme','xiaomi','redmi','poco','mi',
              'tecno','infinix','honor','nokia','asus','lava','micromax','lenovo','sony','lg','huawei',
              'iphone','google','pixel','nothing']:
        if norm.startswith(p + ' '):
            norm = norm[len(p)+1:]
            break
    tokens = norm.split()
    keep_words = {'pro','plus','max','ultra','mini','play','note','lite','prime','power','fusion','neo','fe','nord','one','active','edge'}
    ids = []
    for t in tokens:
        if any(c.isdigit() for c in t) or t in keep_words:
            ids.append(t)
    return ids

def find_matched_device(brand, model_phrase):
    """Re-run the same lookup logic the regenerator used. Returns (device_name, codes) or (None, [])."""
    if not brand or not model_phrase: return None, []
    bs = BRAND_ALIASES.get(brand.lower(), brand.lower())
    devices = mongo_lookup.get(bs, [])
    if not devices: return None, []
    target = normalize(model_phrase)
    target_no_brand = target
    for p in ['samsung','apple','motorola','oneplus','oppo','vivo','realme','xiaomi','redmi','poco','mi',
              'tecno','infinix','honor','nokia','asus','lava','micromax','lenovo','sony','lg','huawei',
              'iphone','google','pixel','nothing']:
        if target_no_brand.startswith(p + ' '):
            target_no_brand = target_no_brand[len(p)+1:]
            break
    if bs == 'apple' and not target_no_brand.startswith('iphone'):
        target_no_brand = 'iphone ' + target_no_brand
    if brand.lower() in ('redmi','poco','mi') and not target_no_brand.startswith(brand.lower()):
        target_no_brand = brand.lower() + ' ' + target_no_brand
    # 1. Exact
    for dn_lower, dn, codes in devices:
        if dn_lower == target_no_brand: return dn, codes
    # 2. Best-overlap match (mirrors regenerator)
    target_words = set(target_no_brand.split())
    best = None
    for dn_lower, dn, codes in devices:
        if re.search(r'\b[A-Z0-9]{4,}\b', dn) and len(dn) > 25: continue
        dn_words = set(normalize(dn).split())
        if not dn_words: continue
        overlap = len(target_words & dn_words)
        if overlap == 0: continue
        if dn_lower == target_no_brand:
            score = (1000, -len(dn_lower))
        elif target_no_brand.startswith(dn_lower) or dn_lower.startswith(target_no_brand):
            score = (500 + overlap, -abs(len(dn_lower) - len(target_no_brand)))
        else:
            score = (overlap, -abs(len(dn_lower) - len(target_no_brand)))
        if best is None or score > best[0]:
            best = (score, dn, codes)
    if best:
        return best[1], best[2]
    return None, []

# ---------- Detect issues ----------
SRC_COL = {h: i for i, h in enumerate(header)}

def has_model_codes(bp1):
    """BP1 has model codes if there's an alphanum token >=4 chars containing both letters and digits
    OR a clear model code pattern after the brand/model name."""
    if not bp1: return False
    # Look for tokens like XT1922, SM-A325F, NE2211, GLUOG, A2172, M1908C3JI, PAMH0001IN
    # Pattern: alphanumeric token with at least one letter AND (digit OR length>=4 of mixed)
    matches = re.findall(r'\b[A-Z][A-Z0-9/-]{3,}\b', bp1)
    return len(matches) > 0

flagged = []  # list of (sku_row, issue_types_list, issue_details_list)

for row in data_rows:
    sku = row[SRC_COL['SKU']] or ''
    title = row[SRC_COL['item_name']] or ''
    bp1 = row[SRC_COL['bullet1']] or ''
    keywords = row[SRC_COL['keywords']] or ''

    issue_types = []
    issue_details = []

    # 1. multi_variant_title
    brand, model_phrase = parse_title(title)
    if '/' in model_phrase:
        issue_types.append('multi_variant_title')
        issue_details.append(f'Title has multi-device variants: "{model_phrase}"')

    # 2. no_model_codes
    if not has_model_codes(bp1):
        issue_types.append('no_model_codes')
        issue_details.append('BP1 has no model code tokens (lookup failed or returned empty)')

    # 3. potential_device_mismatch — only if model codes were found
    if has_model_codes(bp1) and brand and model_phrase:
        matched_dn, matched_codes = find_matched_device(brand, model_phrase)
        if matched_dn:
            id_tokens = extract_id_tokens(model_phrase)
            matched_norm = normalize(matched_dn)
            missing = [t for t in id_tokens if t not in matched_norm and t not in matched_norm.replace(' ', '')]
            # Also check if numeric ID like "g3" is present as part of "g (3rd gen)"
            for t in list(missing):
                if any(c.isdigit() for c in t):
                    digit = re.search(r'\d+', t).group(0)
                    if digit in matched_norm:
                        missing.remove(t)
            if missing:
                issue_types.append('potential_device_mismatch')
                issue_details.append(f'Title model "{model_phrase}" missing {missing} in matched device "{matched_dn}"')

    # 4. bare_gold_in_keywords
    if keywords and re.search(r'\bgold\b', keywords, re.IGNORECASE):
        # If keywords has gold AND title doesn't have (Gold), flag
        if not re.search(r'\(gold\)', title, re.IGNORECASE):
            pass  # title has no gold either, skip
        else:
            # Title has (Gold) but keywords has bare gold — minor issue
            issue_types.append('bare_gold_in_keywords')
            issue_details.append('keywords field has "gold" without parens')

    if issue_types:
        flagged.append((row, issue_types, issue_details))

print(f'\nFlagged {len(flagged)} SKUs out of {len(data_rows)}')
print('Issue type distribution:')
from collections import Counter
issue_counter = Counter()
for _, types, _ in flagged:
    for t in types:
        issue_counter[t] += 1
for t, c in issue_counter.most_common():
    print(f'  {t}: {c}')

# ---------- Write v5 with both sheets ----------
print(f'\nWriting {DEST}...')
dst_wb = Workbook()

# Sheet 1: Template (full data)
ws1 = dst_wb.active
ws1.title = 'Template'
ws1.append(header)
for row in data_rows:
    ws1.append(list(row))

# Sheet 2: Flagged (problematic + issue columns)
ws2 = dst_wb.create_sheet('Flagged')
flagged_header = list(header) + ['issue_types', 'issue_details']
ws2.append(flagged_header)

# Style the header
hdr_fill = PatternFill('solid', fgColor='FFD966')
hdr_font = Font(bold=True)
for cell in ws2[1]:
    cell.fill = hdr_fill
    cell.font = hdr_font

# Color-code issue rows by severity
COLOR_BY_ISSUE = {
    'no_model_codes':            'FFE699',  # yellow — minor (still uploadable, just less searchable)
    'potential_device_mismatch': 'F4B084',  # orange — accuracy concern
    'multi_variant_title':       'FFE699',  # yellow — explains why no codes
    'bare_gold_in_keywords':     'BDD7EE',  # light blue — cosmetic
}

for sku_row, types, details in flagged:
    out = list(sku_row) + [', '.join(types), ' | '.join(details)]
    ws2.append(out)
    # Apply color based on most-severe issue
    last_row = ws2.max_row
    severity_order = ['potential_device_mismatch', 'no_model_codes', 'multi_variant_title', 'bare_gold_in_keywords']
    color = None
    for s in severity_order:
        if s in types:
            color = COLOR_BY_ISSUE.get(s)
            break
    if color:
        fill = PatternFill('solid', fgColor=color)
        for cell in ws2[last_row]:
            cell.fill = fill

# Set column widths for Flagged sheet
for col_letter, width in [('A', 18), ('D', 50), ('P', 60), ('AC', 30), ('AD', 80)]:
    try: ws2.column_dimensions[col_letter].width = width
    except: pass

dst_wb.save(DEST)
print(f'Wrote {DEST}')
print(f'  Sheet 1: Template — {len(data_rows)} SKUs')
print(f'  Sheet 2: Flagged  — {len(flagged)} SKUs (color-coded by issue type)')
