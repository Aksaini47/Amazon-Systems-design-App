import openpyxl
import re

# Load workbook
print("Loading PHONE_ACCESSORY_FILLED.xlsm...")
wb = openpyxl.load_workbook('PHONE_ACCESSORY_FILLED.xlsm', data_only=True)
ws = wb['Template']

# Column indices (1-based)
COL_SKU = 1
COL_ITEM_NAME = 7
COL_BULLET1 = 31
COL_BULLET2 = 32
COL_BULLET3 = 33
COL_BULLET4 = 34
COL_BULLET5 = 35
COL_KEYWORD = 36

def detect_tier(title_str):
    """Detect tier: Bronze (LCD), Silver (OLED), Gold (CareOG)"""
    if 'CareOG' in title_str:
        return 'gold'
    elif 'OLED' in title_str or 'Amoled' in title_str:
        return 'silver'
    elif 'LCD' in title_str:
        return 'bronze'
    else:
        return 'bronze'  # default fallback

def detect_fingerprint(title_str):
    """Detect fingerprint status"""
    # Check for explicit fingerprint support
    if 'Fingerprint Support)' in title_str and 'No Fingerprint Support)' not in title_str:
        return 'Fingerprint Supported'
    # Check for "No Fingerprint Support" - this is non fingerprint
    if 'No Fingerprint Support' in title_str:
        return 'Non Fingerprint'
    # OLED/AMOLED always support fingerprint
    if 'OLED' in title_str or 'Amoled' in title_str:
        return 'Fingerprint Supported'
    return 'Non Fingerprint'

def detect_frame(title_str):
    """Detect if with frame or without frame - ONLY if explicitly mentioned"""
    # Case insensitive check for With Frame
    if 'with frame' in title_str.lower() or ' WF' in title_str or title_str.endswith('WF'):
        return 'With Frame'
    return 'Without Frame'

def extract_model_info(title):
    """Extract device model from item name title."""
    if not title:
        return None, None

    title_str = str(title)

    # Extract brand and model - pattern "Compatible for [Brand] [Model] ..."
    match = re.search(r'Compatible for\s+([A-Za-z0-9\-\s]+?)(?:\s+(?:LCD|OLED|Display|CareOG|Amoled)|$)', title_str)
    if match:
        device_str = match.group(1).strip()
        # Remove parenthetical content
        device_str = re.sub(r'\s*\([^)]*\)\s*', ' ', device_str).strip()
        parts = device_str.split(' ', 1)
        if len(parts) >= 2:
            brand = parts[0]
            model = parts[1].strip()
        else:
            brand = parts[0]
            model = parts[0]
    else:
        # Fallback
        fallback = re.sub(r'Compatible for\s+', '', title_str)
        fallback = re.split(r'\s+(?:LCD|OLED|Display|CareOG|Amoled|Display)', fallback)[0].strip()
        fallback = re.sub(r'\s*\([^)]*\)\s*', ' ', fallback).strip()
        parts = fallback.split(' ', 1)
        brand = parts[0] if parts else None
        model = parts[1] if len(parts) > 1 else fallback

    return brand, model

def generate_bullet1(brand, model, fingerprint):
    """Bullet 1: Compatibility + fingerprint note"""
    if not model:
        return ""

    if brand and brand.lower() in ['iphone', 'apple']:
        return f"Compatible with {brand} {model} — Check Settings > General > About for exact Model Number before ordering"
    elif brand and brand.lower() in ['samsung']:
        return f"Compatible with Samsung {model} — Check Settings > About Phone for exact Model Number before ordering"
    else:
        fp_note = " — Fingerprint not supported on this device" if fingerprint == 'Non Fingerprint' else " — Fingerprint supported after fitting"
        return f"Compatible with {brand} {model}{fp_note} — Verify your model number from Settings before ordering"

def generate_bullet2(tier, fingerprint, frame):
    """Bullet 2: Tier value proposition + fingerprint clarity"""
    if tier == 'gold':
        return "Gold Tier CareOG — Tested OG combo with premium quality display and touch. Superior reliability for professional repair technicians."
    elif tier == 'silver':
        if fingerprint == 'Fingerprint Supported':
            return "Silver Tier OLED — Premium display quality with fingerprint sensor support. Vivid colors and better contrast than LCD."
        else:
            return "Silver Tier OLED — Premium display quality with vivid colors and better contrast. Great value for price."
    else:  # bronze
        return "Bronze Tier LCD — Budget friendly good quality display. Great value for money, reliable performance for everyday use."

def generate_bullet3(tier, frame):
    """Bullet 3: Quality + frame details"""
    if tier == 'gold':
        if frame == 'With Frame':
            return "Quality Tested — 100% checked for dead pixels, color calibration, and touch response. With frame for easy installation."
        else:
            return "Quality Tested — 100% checked for dead pixels, color calibration, and touch response. Screen only, frame not included."
    elif tier == 'silver':
        if frame == 'With Frame':
            return "OLED Quality — Original chip IC, vibrant colors, tested panel. With frame for perfect fit and easy installation."
        else:
            return "OLED Quality — Original chip IC, vibrant colors, tested panel. Screen only, professional installation recommended."
    else:  # bronze
        if frame == 'With Frame':
            return "LCD Quality — Good color reproduction, original chip compatible, tested before dispatch. With frame for easy installation."
        else:
            return "LCD Quality — Good color reproduction, original chip compatible, tested before dispatch. Screen only, frame not included."

def generate_bullet4(frame, fingerprint):
    """Bullet 4: Build variant + installation"""
    if frame == 'With Frame':
        if fingerprint == 'Fingerprint Supported':
            return "With Frame Assembly — Pre-framed screen, ready to install. Fingerprint sensor requires framing during fitting. Adhesive pre-applied."
        else:
            return "With Frame Assembly — Screen pre-pasted on frame, ready to install. Original adhesive pre-applied. Colored frame rims shipped based on availability."
    else:
        return "Screen Only — Requires frame transfer from your old screen. Adhesive sheet included. Professional installation recommended."

def generate_bullet5():
    """Bullet 5: Warranty + support"""
    return "Warranty: 7 days replacement — QC tested before dispatch. No warranty after protection film removal or installation. For help visit Repairfully.com"

def generate_keywords(brand, tier, fingerprint):
    """Generate backend keywords (200 bytes max)"""
    # Base keywords
    base = "display sceen replacement mobile repair"

    # Add tier keywords
    if tier == 'gold':
        tier_kw = "careog tested combo premium quality"
    elif tier == 'silver':
        tier_kw = "oled amoled premium fingerprint"
    else:
        tier_kw = "lcd budget friendly cheap"

    # Add fingerprint
    if fingerprint == 'Fingerprint Supported':
        fp_kw = "fingerprint support"
    else:
        fp_kw = "non fingerprint"

    # Add brand
    if brand and brand.lower() in ['iphone', 'apple']:
        brand_kw = "apple iphone"
    elif brand and brand.lower() in ['samsung']:
        brand_kw = "samsung galaxy"
    else:
        brand_kw = ""

    keywords = f"{base} {tier_kw} {fp_kw} {brand_kw}".strip()
    return keywords

# Process all rows
print("\nProcessing listings...")
filled_count = 0
errors = []

for row in range(7, ws.max_row + 1):
    sku = ws.cell(row, COL_SKU).value
    title = ws.cell(row, COL_ITEM_NAME).value

    if not title or not str(title).strip():
        continue

    try:
        title_str = str(title)

        # Detect attributes
        tier = detect_tier(title_str)
        fingerprint = detect_fingerprint(title_str)
        frame = detect_frame(title_str)

        # Extract model info
        brand, model = extract_model_info(title)

        # Generate bullets
        bullet1 = generate_bullet1(brand, model, fingerprint)
        bullet2 = generate_bullet2(tier, fingerprint, frame)
        bullet3 = generate_bullet3(tier, frame)
        bullet4 = generate_bullet4(frame, fingerprint)
        bullet5 = generate_bullet5()

        # Generate keywords
        keywords = generate_keywords(brand, tier, fingerprint)

        # Write to cells
        ws.cell(row, COL_BULLET1).value = bullet1
        ws.cell(row, COL_BULLET2).value = bullet2
        ws.cell(row, COL_BULLET3).value = bullet3
        ws.cell(row, COL_BULLET4).value = bullet4
        ws.cell(row, COL_BULLET5).value = bullet5
        ws.cell(row, COL_KEYWORD).value = keywords

        filled_count += 1

    except Exception as e:
        errors.append(f"Row {row} ({sku}): {str(e)}")

print(f"\nFilled: {filled_count} listings")
if errors:
    print(f"Errors: {len(errors)}")

# Save
output_file = 'PHONE_ACCESSORY_FILLED_bullets_v2.xlsm'
print(f"Saving to {output_file}...")
wb.save(output_file)
wb.close()
print("Done!")

# Print samples for each tier
print(f"\n{'='*60}")
print("SAMPLE OUTPUT BY TIER:")
print(f"{'='*60}")

wb2 = openpyxl.load_workbook(output_file, data_only=True)
ws2 = wb2['Template']

for tier_name, search_term in [('Gold (CareOG)', 'CareOG'), ('Silver (OLED)', 'OLED'), ('Bronze (LCD)', 'LCD')]:
    print(f"\n--- {tier_name} ---")
    for row in range(7, ws2.max_row + 1):
        title = ws2.cell(row, COL_ITEM_NAME).value
        if title and search_term in str(title) and 'CareOG' in str(title) if tier_name == 'Gold' else (search_term in str(title) and 'CareOG' not in str(title)):
            if tier_name == 'Gold' and 'CareOG' not in str(title):
                continue
            print(f"Title: {str(title)[:70]}...")
            print(f"B1: {ws2.cell(row, COL_BULLET1).value}")
            print(f"B2: {ws2.cell(row, COL_BULLET2).value}")
            print(f"B3: {ws2.cell(row, COL_BULLET3).value}")
            print(f"B4: {ws2.cell(row, COL_BULLET4).value}")
            print(f"B5: {ws2.cell(row, COL_BULLET5).value}")
            print(f"KW: {ws2.cell(row, COL_KEYWORD).value}")
            break

wb2.close()