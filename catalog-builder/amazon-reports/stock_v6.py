import re
from collections import defaultdict
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# STEP 1: Parse ACTUAL STOCK
# ============================================================
with open('actual stock.txt', 'r', encoding='utf-8') as f:
    slines = f.readlines()

BRAND_KEYWORDS = ['REALME', 'VIVO', 'OPPO', 'ONEPLUS', 'APPLE', 'SAMSUNG',
                  'XIAOMI', 'REDMI', 'POCO', 'MIX', 'MOTO', 'HONOR',
                  'NOKIA', 'ASUS', 'NOTHING', 'INFINIX']

stock = []
current_brand = ''
for line in slines:
    stripped = line.rstrip().strip()
    if not stripped or stripped.lower() == 'actual stock':
        continue
    is_brand = stripped.isupper() and any(b in stripped for b in BRAND_KEYWORDS)
    if is_brand:
        current_brand = stripped
    else:
        name = stripped
        if name and current_brand:
            qty_match = re.search(r' - (\d+)$', name)
            qty = int(qty_match.group(1)) if qty_match else 1
            clean_name = re.sub(r' - \d+$', '', name).strip()
            nl = clean_name.lower()

            if 'oled' in nl or 'amoled' in nl:
                ss = 'OLED'
            elif 'incell' in nl:
                ss = 'Incell LCD'
            elif 'tft' in nl:
                ss = 'TFT LCD'
            elif 'lcd' in nl:
                ss = 'LCD'
            else:
                ss = 'Not Specified'

            if 'careog' in nl:
                sq = 'CareOG'
            elif 'frame' in nl or 'with frame' in nl or ' wf ' in nl:
                sq = 'With Frame'
            else:
                sq = 'Standard'

            stock.append({
                'brand': current_brand,
                'original_name': clean_name,
                'qty': qty,
                'screen': ss,
                'quality': sq
            })

# ============================================================
# STEP 2: Parse AMAZON LISTINGS
# ============================================================
with open('All+Listings+Report_05-12-2026.txt', 'r', encoding='utf-8-sig') as f:
    lines = f.readlines()

listings = []
for l in lines[1:]:
    if not l.strip():
        continue
    cols = [c.strip() for c in l.split('\t')]
    if len(cols) < 30:
        cols += [''] * (30 - len(cols))
    listings.append({
        'title': cols[0], 'sku': cols[3],
        'price': cols[4], 'quantity': cols[5], 'status': cols[29]
    })

# ============================================================
# STEP 3: Extract from Amazon listing title
# ============================================================
def extract_compatible_model(title):
    if not title or 'Compatible for' not in title:
        return None, None
    after = title.split('Compatible for', 1)[1].strip()
    segment = after.split('/')[0].strip()
    segment = re.sub(r'\([^)]*\)', '', segment)
    for skip in ['CareOG', 'OLED', 'LCD', 'Incell', 'TFT', 'Amoled', 'AMOLED',
                 'Super OLED', 'Fingerprint Support', 'No Fingerprint Support',
                 'Display+Touch Screen Combo', 'Display Screen Replacement Combo',
                 'Display Screen Combo', 'Screen Combo', 'Display Touch Digitizer',
                 'Assembly', 'Replacement', 'with frame', 'Standard', 'White',
                 'Black', 'Gold', 'Folder']:
        segment = re.sub(re.escape(skip), ' ', segment, flags=re.IGNORECASE)
    segment = re.sub(r'\s+', ' ', segment).strip()

    # Strip brand prefix so keys align with stock model keys
    for strip in ['Apple iPhone ', 'Samsung Galaxy ', 'Samsung ', 'Xiaomi ',
                  'Redmi ', 'POCO ', 'Realme ', 'Vivo ', 'Oppo ', 'OnePlus ',
                  'Motorola Moto ', 'Moto ', 'Honor ', 'Nokia ', 'Asus ',
                  'Nothing Phone ', 'Nothing ', 'Infinix ']:
        if segment.lower().startswith(strip.lower()):
            segment = segment[len(strip):].strip()
            break

    brand = None
    checks = [
        ('Apple iPhone', 'APPLE'), ('Samsung Galaxy', 'SAMSUNG'),
        ('Samsung ', 'SAMSUNG'), ('Xiaomi ', 'XIAOMI'), ('Redmi ', 'REDMI'),
        ('POCO ', 'POCO'), ('Realme ', 'REALME'), ('Vivo ', 'VIVO'),
        ('Oppo ', 'OPPO'), ('OnePlus ', 'ONEPLUS'), ('Motorola Moto', 'MOTO'),
        ('Moto ', 'MOTO'), ('Honor ', 'HONOR'), ('Nokia ', 'NOKIA'),
        ('Asus ', 'ASUS'), ('Nothing Phone', 'NOTHING'), ('Nothing ', 'NOTHING'),
        ('Infinix ', 'INFINIX'),
    ]
    for kw, br in checks:
        if kw.lower() in after.lower():
            brand = br
            break
    return segment, brand


def extract_screen_quality(title):
    t = title.lower()
    if 'super oled' in t or 'amoled' in t or ('oled' in t and 'lcd' not in t and 'incell' not in t):
        ls = 'OLED'
    elif 'incell' in t:
        ls = 'Incell LCD'
    elif 'tft' in t:
        ls = 'TFT LCD'
    elif 'lcd' in t:
        ls = 'LCD'
    else:
        ls = 'Not Specified'

    if 'careog' in t:
        lq = 'CareOG'
    elif 'with frame' in t or '(with frame)' in t or ' wf ' in t:
        lq = 'With Frame'
    else:
        lq = 'Standard'
    return ls, lq


listing_by_model = defaultdict(list)
for row in listings:
    title = row['title']
    if not title:
        continue
    model_name, model_brand = extract_compatible_model(title)
    if model_name:
        ls, lq = extract_screen_quality(title)
        listing_by_model[model_name].append({
            'row': row, 'model_name': model_name, 'brand': model_brand,
            'listing_screen': ls, 'listing_quality': lq
        })

# ============================================================
# STEP 4: BRAND NORMALIZATION — Sunsky -> Stock
# ============================================================
SUNSKY_BRAND_MAP = {
    # Sunsky brand name (lowercase) -> stock brand constant
    'apple': 'APPLE',
    'samsung': 'SAMSUNG',
    'oneplus': 'ONEPLUS',
    'oppo': 'OPPO',
    'vivo': 'VIVO',
    'realme': 'REALME',
    'xiaomi': 'XIAOMI',
    'redmi': 'XIAOMI',
    'poco': 'XIAOMI',
    'motorola': 'MOTO',
    'honor': 'HONOR',
    'nokia': 'NOKIA',
    'asus': 'ASUS',
    'nothing': 'NOTHING',
    'infinix': 'INFINIX',
    'tecno': 'MIX',
    'itel': 'MIX',
    'iqoo': 'XIAOMI',
}

# ============================================================
# STEP 5: Parse SUNSKY — BRAND-AWARE INDEX
# ============================================================
sunsky_df = pd.read_excel(
    r'C:\Users\DELL\Claude Repairfully.com\Sunsky Products data\AllImages_cleaned\Sunsky_MatchedProducts_R41_20260508.xlsx'
)

# Build brand-aware indices:
# sunsky_by_brand[stock_brand][model_key] = list of entries
# sunsky_crossfit_by_brand[stock_brand][model_key] = list of entries (from cross-fit models)
sunsky_by_brand = defaultdict(lambda: defaultdict(list))
sunsky_crossfit_by_brand = defaultdict(lambda: defaultdict(list))

for i, row in sunsky_df.iterrows():
    sunsky_brand_raw = str(row['Compatible Brand']).strip() if pd.notna(row['Compatible Brand']) else ''
    sunsky_brand = sunsky_brand_raw.lower()
    stock_brand = SUNSKY_BRAND_MAP.get(sunsky_brand, None)

    primary = str(row['Primary Model']).strip() if pd.notna(row['Primary Model']) else ''
    crossfit = str(row['Cross-fit Models']).strip() if pd.notna(row['Cross-fit Models']) else ''
    item_no = str(row['Item Number']).strip() if pd.notna(row['Item Number']) else ''
    price_inr = str(row['Price (INR)']).strip() if pd.notna(row['Price (INR)']) else ''
    product_type = str(row['Product Type']).strip() if pd.notna(row['Product Type']) else ''

    if not primary or not stock_brand:
        continue

    pkey = re.sub(r'[^a-z0-9]', '', primary.lower())
    if len(pkey) < 1:
        continue

    entry = {
        'sunsky_brand': sunsky_brand_raw,
        'primary': primary,
        'crossfit': crossfit,
        'item_no': item_no,
        'price_inr': price_inr,
        'product_type': product_type,
        'category': str(row['Category']).strip() if pd.notna(row.get('Category')) else ''
    }

    # Primary model index
    sunsky_by_brand[stock_brand][pkey].append(entry)

    # Cross-fit models index (also brand-aware)
    if crossfit:
        for cf in crossfit.split('/'):
            cf = cf.strip()
            if not cf:
                continue
            cfkey = re.sub(r'[^a-z0-9]', '', cf.lower())
            if len(cfkey) >= 1 and cfkey != pkey:
                sunsky_crossfit_by_brand[stock_brand][cfkey].append(entry)

# ============================================================
# STEP 5B: Build DISPLAY ASSEMBLY + FULL_DEVICE_NAME index
# Stock items are mobile displays
# Match against Full_Device_Name column (skip brand/model matching)
# ============================================================
sunsky_display_devices = {}  # normalized_full_device_name -> entry

for i, row in sunsky_df.iterrows():
    category = str(row['Category']).strip() if pd.notna(row.get('Category')) else ''

    # ONLY index Display Assembly category
    if 'display assembly' not in category.lower():
        continue

    full_device = str(row['Full_Device_Name']).strip() if pd.notna(row.get('Full_Device_Name')) else ''
    if not full_device or full_device == 'nan':
        continue

    crossfit = str(row['Cross-fit Models']).strip() if pd.notna(row['Cross-fit Models']) else ''
    item_no = str(row['Item Number']).strip() if pd.notna(row['Item Number']) else ''
    price_inr = str(row['Price (INR)']).strip() if pd.notna(row['Price (INR)']) else ''
    product_type = str(row['Product Type']).strip() if pd.notna(row['Product Type']) else ''

    # Normalize Full_Device_Name - keep brand prefix for matching
    fd_lower = full_device.lower()
    for prefix in ['apple iphone ', 'samsung galaxy ', 'samsung ', 'xiaomi ',
                   'redmi ', 'poco ', 'realme ', 'vivo ', 'oppo ', 'oneplus ',
                   'motorola moto ', 'moto ', 'honor ', 'nokia ', 'asus ',
                   'nothing phone ', 'nothing ', 'infinix ']:
        if fd_lower.startswith(prefix):
            # Keep brand + model, just normalize spaces
            fd_normalized = (prefix.strip() + ' ' + fd_lower[len(prefix):].strip()).strip()
            break
    else:
        fd_normalized = fd_lower

    # Clean up for matching - keep brand
    fd_key = re.sub(r'[^a-z0-9]', '', fd_normalized)
    if len(fd_key) < 1:
        continue

    entry = {
        'full_device_name': full_device,
        'crossfit': crossfit,
        'item_no': item_no,
        'price_inr': price_inr,
        'product_type': product_type,
        'category': category
    }

    # If multiple entries for same device, keep one with crossfit if available
    if fd_key not in sunsky_display_devices or (not sunsky_display_devices[fd_key]['crossfit'] and crossfit):
        sunsky_display_devices[fd_key] = entry

# Also build cross-fit index
sunsky_display_crossfits = defaultdict(list)
for fd_key, entry in sunsky_display_devices.items():
    if entry['crossfit']:
        for cf in entry['crossfit'].split('/'):
            cf = cf.strip()
            if not cf:
                continue
            cf_key = re.sub(r'[^a-z0-9]', '', cf.lower())
            if cf_key and cf_key != fd_key and len(cf_key) >= 2:
                sunsky_display_crossfits[cf_key].append(entry)

# ============================================================
# STEP 6: Stock model normalization
# ============================================================
BRAND_PREFIX_MAP = {
    'APPLE': ['Apple iPhone '],
    'SAMSUNG': ['Samsung Galaxy '],
    'ONEPLUS': ['OnePlus '],
    'OPPO': ['Oppo '],
    'VIVO': ['Vivo '],
    'REALME': ['Realme '],
    'XIAOMI': ['Xiaomi ', 'Redmi ', 'POCO '],
    'REDMI': ['Xiaomi ', 'Redmi ', 'POCO '],
    'POCO': ['Xiaomi ', 'Redmi ', 'POCO '],
    'MOTO': ['Moto '],
    'HONOR': ['Honor '],
    'NOKIA': ['Nokia '],
    'ASUS': ['Asus '],
    'NOTHING': ['Nothing Phone ', 'Nothing '],
    'INFINIX': ['Infinix '],
    'MIX': ['Infinix ', 'Honor ', 'Nokia ', 'Moto ', 'Asus ', 'Nothing Phone ', 'Nothing '],
}

def get_stock_model(name, brand):
    n = name
    n = re.sub(r'^iPhone\s+', 'Apple iPhone ', n, flags=re.IGNORECASE)
    n = re.sub(r'^Galaxy\s+', 'Samsung Galaxy ', n, flags=re.IGNORECASE)
    prefixes = BRAND_PREFIX_MAP.get(brand, [])
    for prefix in prefixes:
        if n.startswith(prefix):
            return prefix, n[len(prefix):].strip()
    return '', n


def normalize_stock_model_full(name, brand):
    prefix, model = get_stock_model(name, brand)
    n = model
    for skip in ['Incell', 'OLED', 'AMOLED', 'Super OLED', 'LCD', 'TFT',
                 'CareOG', 'Frame', 'With Frame', 'Standard', 'White', 'Black', 'Gold']:
        n = re.sub(r'\b' + re.escape(skip) + r'\b', '', n, flags=re.IGNORECASE)
    n = re.sub(r'\([^)]*\)', '', n)
    n = re.sub(r'\s+', ' ', n).strip()
    return (prefix + n).strip()


def get_model_num_key(name, brand):
    _, model = get_stock_model(name, brand)
    n = model
    for skip in ['Incell', 'OLED', 'AMOLED', 'Super OLED', 'LCD', 'TFT',
                 'CareOG', 'Frame', 'With Frame', 'Standard', 'White', 'Black', 'Gold']:
        n = re.sub(r'\b' + re.escape(skip) + r'\b', '', n, flags=re.IGNORECASE)
    n = re.sub(r'\([^)]*\)', '', n)
    n = re.sub(r'\s+', ' ', n).strip()
    return re.sub(r'[^a-z0-9]', '', n.lower())


def get_full_device_key(name, brand):
    """
    Get full device key including brand prefix for matching against Full_Device_Name.
    E.g., 'Samsung Galaxy A73 OLED' -> 'samsunggalaxya73'
    """
    full_name = name.lower()
    # Normalize iPhone and Galaxy prefixes
    full_name = re.sub(r'^iphone\s+', 'apple iphone ', full_name, flags=re.IGNORECASE)
    full_name = re.sub(r'^galaxy\s+', 'samsung galaxy ', full_name, flags=re.IGNORECASE)

    # Remove quality/screen terms
    for skip in ['incell', 'oled', 'amoled', 'super oled', 'lcd', 'tft',
                 'careog', 'frame', 'with frame', 'standard', 'white', 'black', 'gold']:
        full_name = re.sub(r'\b' + skip + r'\b', '', full_name, flags=re.IGNORECASE)
    full_name = re.sub(r'\([^)]*\)', '', full_name)
    full_name = re.sub(r'\s+', ' ', full_name).strip()

    return re.sub(r'[^a-z0-9]', '', full_name)


def get_display_brand(sbrand, sname):
    """Get display brand - resolve MIX to sub-brand, then map to display name."""
    actual = get_actual_brand(sbrand, sname)
    return BRAND_DISPLAY.get(actual, actual)

BRAND_DISPLAY = {
    'APPLE': 'Apple', 'SAMSUNG': 'Samsung', 'ONEPLUS': 'OnePlus',
    'OPPO': 'Oppo', 'VIVO': 'Vivo', 'REALME': 'Realme',
    'XIAOMI': 'Xiaomi/Redmi/Poco', 'MOTO': 'Motorola',
    'HONOR': 'Honor', 'NOKIA': 'Nokia', 'ASUS': 'Asus',
    'INFINIX': 'Infinix', 'MIX': 'Mixed', 'NOTHING': 'Nothing',
    'REDMI': 'Redmi', 'POCO': 'Poco'
}

BRAND_ALIASES = {
    'XIAOMI': ['XIAOMI', 'REDMI', 'POCO'],
    'REDMI': ['XIAOMI', 'REDMI', 'POCO'],
    'POCO': ['XIAOMI', 'REDMI', 'POCO'],
    'MIX': ['INFINIX', 'HONOR', 'NOKIA', 'MOTO', 'ASUS', 'NOTHING'],
}

# Sub-brand detection for MIX stock brand
# When stock brand is MIX, infer actual sub-brand from item name
MIX_SUB_BRAND_DETECT = [
    (r'\bhonor\b', 'HONOR'),
    (r'\bnokia\b', 'NOKIA'),
    (r'\bmoto\b|\bmotorola\b', 'MOTO'),
    (r'\basus\b', 'ASUS'),
    (r'\bnothing\b', 'NOTHING'),
    (r'\binfinix\b', 'INFINIX'),
]

def get_actual_brand(stock_brand, stock_name):
    """Resolve MIX to actual sub-brand, return stock brand for others."""
    if stock_brand != 'MIX':
        return stock_brand
    n = stock_name.lower()
    for pattern, resolved in MIX_SUB_BRAND_DETECT:
        if re.search(pattern, n):
            return resolved
    return 'MIX'  # fallback, keep as MIX


def screen_match(stock_screen, listing_screen):
    s = stock_screen
    l = listing_screen
    if s == l:
        return True
    if s in ['Incell LCD', 'LCD'] and l in ['Incell LCD', 'LCD']:
        return True
    return False


def get_flags(stock_screen, stock_quality, listing_screen, listing_quality):
    flags = []
    if not screen_match(stock_screen, listing_screen):
        flags.append(f'SCREEN: {stock_screen}->{listing_screen}')
    if stock_quality != listing_quality:
        flags.append(f'QUALITY: {stock_quality}->{listing_quality}')
    return ' | '.join(flags)


def find_amazon_match(stock_item):
    sbrand = get_actual_brand(stock_item['brand'], stock_item['original_name'])
    sname = stock_item['original_name']
    ss = stock_item['screen']
    sq = stock_item['quality']

    norm_model = normalize_stock_model_full(sname, sbrand)
    model_num = get_model_num_key(sname, sbrand)

    exact = []
    loose = []
    seen = set()

    if norm_model in listing_by_model:
        for item in listing_by_model[norm_model]:
            sku = item['row']['sku']
            if sku in seen:
                continue
            seen.add(sku)
            lb = item['brand']
            if lb and lb != sbrand and lb not in BRAND_ALIASES.get(sbrand, [sbrand]):
                continue
            if screen_match(ss, item['listing_screen']) and item['listing_quality'] == sq:
                exact.append(item)
            else:
                loose.append(item)

    if model_num and len(model_num) >= 1:
        for listing_model, items in listing_by_model.items():
            listing_key = re.sub(r'[^a-z0-9]', '', listing_model.lower())
            if model_num in listing_key:
                for item in items:
                    sku = item['row']['sku']
                    if sku in seen:
                        continue
                    seen.add(sku)
                    lb = item['brand']
                    if lb and lb != sbrand and lb not in BRAND_ALIASES.get(sbrand, [sbrand]):
                        continue
                    if screen_match(ss, item['listing_screen']) and item['listing_quality'] == sq:
                        exact.append(item)
                    else:
                        loose.append(item)

    exact_active = [c for c in exact if c['row']['status'] == 'Active']
    loose_active = [c for c in loose if c['row']['status'] == 'Active']

    if exact_active:
        return exact_active[0], 'EXACT'
    if exact:
        return exact[0], 'EXACT'
    if loose_active:
        return loose_active[0], 'LOOSE'
    if loose:
        return loose[0], 'LOOSE'
    return None, 'NOT LISTED'


def find_sunsky_match(stock_name, stock_brand, stock_quality=''):
    """
    Match stock item against Sunsky Full_Device_Name in Display Assembly category.
    ONLY exact matches on full device name (brand + model).
    """
    # Use full device key (with brand) for matching against Full_Device_Name
    model_num = get_full_device_key(stock_name, stock_brand)

    if not model_num or len(model_num) < 1:
        return None, None

    # Exact match only
    if model_num in sunsky_display_devices:
        return sunsky_display_devices[model_num], 'display_exact'

    return None, None


# ============================================================
# STEP 7: Build rows for all 3 categories
# ============================================================
exact_rows = []
loose_rows = []
not_listed_rows = []

for s in stock:
    amazon_match, match_type = find_amazon_match(s)
    sunsky_match, sunsky_type = find_sunsky_match(s['original_name'], s['brand'], s['quality'])

    if sunsky_match:
        sunsky_full_device_name = sunsky_match['full_device_name']
        sunsky_crossfit = sunsky_match['crossfit']
        sunsky_item_no = sunsky_match['item_no']
        sunsky_price_inr = sunsky_match['price_inr']
        sunsky_product_type = sunsky_match['product_type']
        sunsky_category = sunsky_match['category']
    else:
        sunsky_full_device_name = ''
        sunsky_crossfit = ''
        sunsky_item_no = ''
        sunsky_price_inr = ''
        sunsky_product_type = ''
        sunsky_category = ''

    base = {
        'Brand': get_display_brand(s['brand'], s['original_name']),
        'Stock Item': s['original_name'],
        'Qty': s['qty'],
        'Stock Screen': s['screen'],
        'Stock Quality': s['quality'],
        'Sunsky Category': sunsky_category,
        'Sunsky Full Device Name': sunsky_full_device_name,
        'Sunsky Cross-fit Models': sunsky_crossfit,
        'Sunsky Item No': sunsky_item_no,
        'Sunsky Price (INR)': sunsky_price_inr,
        'Sunsky Product Type': sunsky_product_type,
    }

    if amazon_match:
        row_data = dict(base)
        row_data.update({
            'Amazon Model Matched': amazon_match['model_name'],
            'Amazon Status': amazon_match['row']['status'],
            'Amazon SKU': amazon_match['row']['sku'],
            'Amazon Qty': amazon_match['row']['quantity'],
            'Amazon Price': amazon_match['row']['price'],
            'Listing Screen': amazon_match['listing_screen'],
            'Listing Quality': amazon_match['listing_quality'],
            'Amazon Title': amazon_match['row']['title'][:90]
        })

        if match_type == 'EXACT':
            exact_rows.append(row_data)
        else:
            row_data['Flag'] = get_flags(
                s['screen'], s['quality'],
                amazon_match['listing_screen'], amazon_match['listing_quality']
            )
            loose_rows.append(row_data)
    else:
        row_data = dict(base)
        row_data.update({
            'Amazon Model Matched': '',
            'Amazon Status': '',
            'Amazon SKU': '',
            'Amazon Qty': '',
            'Amazon Price': '',
            'Listing Screen': '',
            'Listing Quality': '',
            'Amazon Title': '',
            'Flag': ''
        })
        not_listed_rows.append(row_data)

# ============================================================
# STEP 8: Write single Excel file with 4 sheets
# ============================================================
exact_units = sum(r['Qty'] for r in exact_rows)
loose_units = sum(r['Qty'] for r in loose_rows)
not_listed_units = sum(r['Qty'] for r in not_listed_rows)

print('=' * 70)
print('STOCK ANALYSIS — BRAND-AWARE SUNSKY MATCHING')
print('=' * 70)
print()
print(f'1. EXACT MATCH  : {len(exact_rows):>3} items, {exact_units:>3} units')
print(f'2. LOOSE MATCH  : {len(loose_rows):>3} items, {loose_units:>3} units')
print(f'3. NOT LISTED   : {len(not_listed_rows):>3} items, {not_listed_units:>3} units')
print(f'TOTAL           : {len(stock):>3} items, {sum(s["qty"] for s in stock):>3} units')
print()

# Count Sunsky matches
sunsky_found = sum(1 for s in stock if find_sunsky_match(s['original_name'], s['brand'])[0] is not None)
print(f'Sunsky Match    : {sunsky_found} items found in Sunsky database')
print(f'Sunsky No Match : {len(stock) - sunsky_found} items not in Sunsky database')

# Verify Samsung A73 fix
test_samsung_a73 = find_sunsky_match('Galaxy A73 OLED', 'SAMSUNG')
print()
if test_samsung_a73[0]:
    print(f'Test: Samsung Galaxy A73 OLED -> Sunsky: {test_samsung_a73[0]["full_device_name"]}')
    print(f'  Category: {test_samsung_a73[0]["category"]}')
    print(f'  Cross-fit: {test_samsung_a73[0]["crossfit"]}')
else:
    print('Test: Samsung Galaxy A73 OLED -> No Sunsky match')

# Build combined sheet
combined_rows = []
for r in exact_rows:
    r2 = dict(r)
    r2['Match Type'] = 'EXACT MATCH'
    combined_rows.append(r2)
for r in loose_rows:
    r2 = dict(r)
    r2['Match Type'] = 'LOOSE MATCH'
    combined_rows.append(r2)
for r in not_listed_rows:
    r2 = dict(r)
    r2['Match Type'] = 'NOT LISTED'
    combined_rows.append(r2)

# Define columns for each sheet type
EXACT_COLS = ['Brand', 'Stock Item', 'Qty', 'Stock Screen', 'Stock Quality',
              'Sunsky Category', 'Sunsky Full Device Name', 'Sunsky Cross-fit Models',
              'Sunsky Item No', 'Sunsky Price (INR)', 'Sunsky Product Type',
              'Amazon Model Matched', 'Amazon Status', 'Amazon SKU', 'Amazon Qty',
              'Amazon Price', 'Listing Screen', 'Listing Quality', 'Amazon Title']

LOOSE_COLS = EXACT_COLS + ['Flag']

NOTLISTED_COLS = ['Brand', 'Stock Item', 'Qty', 'Stock Screen', 'Stock Quality',
                  'Sunsky Category', 'Sunsky Full Device Name', 'Sunsky Cross-fit Models',
                  'Sunsky Item No', 'Sunsky Price (INR)', 'Sunsky Product Type',
                  'Amazon Model Matched', 'Amazon Status', 'Amazon SKU', 'Amazon Qty',
                  'Amazon Price', 'Listing Screen', 'Listing Quality', 'Amazon Title', 'Flag']

COMBINED_COLS = ['Match Type'] + NOTLISTED_COLS

# Sheet colors
SHEET_COLORS = {
    'Exact Match': 'C6EFCE',  # green
    'Loose Match': 'FFEB9C',  # yellow
    'Not Listed': 'FFC7CE',  # red
    'Combined':    'DDEBF7',  # blue
}

wb = openpyxl.Workbook()

def add_sheet(wb, name, rows, cols):
    ws = wb.create_sheet(title=name)
    fill_color = SHEET_COLORS.get(name, 'FFFFFF')
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    header_font = Font(bold=True)

    # Write header
    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Write data
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(cols, 1):
            val = row_data.get(col_name, '')
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-fit columns
    for col_idx in range(1, len(cols) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    return ws

# ============================================================
# STEP 8: Build Sheet 2 - All Amazon Listings with Cross-fits
# ============================================================
amazon_listing_rows = []

# Function to extract model from title
def extract_model_from_title(title):
    if not title or 'Compatible for' not in title:
        return None, None
    after = title.split('Compatible for', 1)[1].strip()
    segment = after.split('/')[0].strip()
    segment = re.sub(r'\([^)]*\)', '', segment)
    for skip in ['CareOG', 'OLED', 'LCD', 'Incell', 'TFT', 'Amoled', 'AMOLED',
                 'Super OLED', 'Fingerprint Support', 'No Fingerprint Support',
                 'Display+Touch Screen Combo', 'Display Screen Replacement Combo',
                 'Display Screen Combo', 'Screen Combo', 'Display Touch Digitizer',
                 'Assembly', 'Replacement', 'with frame', 'Standard', 'White',
                 'Black', 'Gold', 'Folder']:
        segment = re.sub(re.escape(skip), ' ', segment, flags=re.IGNORECASE)
    segment = re.sub(r'\s+', ' ', segment).strip()

    # Extract brand
    brand = None
    for kw, br in [('Apple iPhone', 'APPLE'), ('Samsung Galaxy', 'SAMSUNG'),
                   ('Samsung ', 'SAMSUNG'), ('Xiaomi ', 'XIAOMI'), ('Redmi ', 'XIAOMI'),
                   ('POCO ', 'XIAOMI'), ('Realme ', 'REALME'), ('Vivo ', 'VIVO'),
                   ('Oppo ', 'OPPO'), ('OnePlus ', 'ONEPLUS'), ('Moto ', 'MOTO'),
                   ('Honor ', 'HONOR'), ('Nokia ', 'NOKIA'), ('Asus ', 'ASUS'),
                   ('Infinix ', 'INFINIX')]:
        if kw.lower() in after.lower():
            brand = br
            break
    return segment, brand

# Function to get full device key
def get_full_device_key_from_title(title):
    model_name, _ = extract_model_from_title(title)
    if not model_name:
        return None
    full_name = model_name.lower()
    full_name = re.sub(r'^iphone\s+', 'apple iphone ', full_name, flags=re.IGNORECASE)
    full_name = re.sub(r'^galaxy\s+', 'samsung galaxy ', full_name, flags=re.IGNORECASE)
    return re.sub(r'[^a-z0-9]', '', full_name)

# Process each listing
for row in listings:
    title = row['title']
    if not title:
        continue

    # Only mobile display listings
    title_lower = title.lower()
    if not any(kw in title_lower for kw in ['display', 'lcd', 'oled', 'amoled', 'touch', 'screen', 'digitizer', 'combo']):
        continue

    model_name, model_brand = extract_model_from_title(title)
    if not model_name:
        continue

    # Find Sunsky match
    full_key = get_full_device_key_from_title(title)
    sunsky_match = None
    if full_key and full_key in sunsky_display_devices:
        sunsky_match = sunsky_display_devices[full_key]
    else:
        # Try partial match
        for key in sunsky_display_devices:
            if len(key) >= 3:
                stock_model = full_key.replace('apple', '').replace('iphone', '').replace('samsung', '').replace('galaxy', '').replace('realme', '').replace('vivo', '').replace('oppo', '').replace('oneplus', '').replace('xiaomi', '').replace('redmi', '').replace('poco', '').replace('honor', '').replace('nokia', '').replace('motorola', '').replace('moto', '').replace('asus', '').replace('infinix', '')
                sunsky_model = key.replace('apple', '').replace('iphone', '').replace('samsung', '').replace('galaxy', '').replace('realme', '').replace('vivo', '').replace('oppo', '').replace('oneplus', '').replace('xiaomi', '').replace('redmi', '').replace('poco', '').replace('honor', '').replace('nokia', '').replace('motorola', '').replace('moto', '').replace('asus', '').replace('infinix', '')
                if len(stock_model) >= 3 and stock_model in sunsky_model:
                    sunsky_match = sunsky_display_devices[key]
                    break

    amazon_listing_rows.append({
        'Amazon SKU': row['sku'],
        'Amazon Title': title[:90],
        'Extracted Model': model_name,
        'Model Brand': model_brand or '',
        'Amazon Status': row['status'],
        'Amazon Price': row['price'],
        'Amazon Qty': row['quantity'],
        'Sunsky Device': sunsky_match['full_device_name'] if sunsky_match else '',
        'Sunsky Cross-fits': sunsky_match['crossfit'] if sunsky_match else '',
        'Sunsky Price (INR)': sunsky_match['price_inr'] if sunsky_match else '',
        'Sunsky Item No': sunsky_match['item_no'] if sunsky_match else '',
    })

print(f'Amazon Listings processed: {len(amazon_listing_rows)}')

# Define Sheet 2 columns
AMAZON_COLS = ['Amazon SKU', 'Amazon Title', 'Extracted Model', 'Model Brand',
               'Amazon Status', 'Amazon Price', 'Amazon Qty',
               'Sunsky Device', 'Sunsky Cross-fits', 'Sunsky Price (INR)', 'Sunsky Item No']

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

# Add sheets
add_sheet(wb, 'Exact Match', exact_rows, EXACT_COLS)
add_sheet(wb, 'Loose Match', loose_rows, LOOSE_COLS)
add_sheet(wb, 'Not Listed', not_listed_rows, NOTLISTED_COLS)
add_sheet(wb, 'Combined', combined_rows, COMBINED_COLS)
add_sheet(wb, 'Amazon Listings + Cross-fits', amazon_listing_rows, AMAZON_COLS)

# Save
output_path = 'stock_analysis_report_v7.xlsx'
wb.save(output_path)

print()
print(f'Excel saved to: amazon-reports/{output_path}')
print('5 sheets: Exact Match | Loose Match | Not Listed | Combined | Amazon Listings + Cross-fits')