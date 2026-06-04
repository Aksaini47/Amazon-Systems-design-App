"""
Copy v6 → v7 with new column appended at end:
  recommended_browse_nodes = "Mobile Displays (11180541031)"   for all 1283 SKUs.

Currently the seller's Category Listings Report shows browse_node = Mobile Screen Guards
(1389425031) — that's the SCREEN PROTECTOR category, not screen replacement parts. Fix it
to the correct leaf node for mobile display assemblies.

Amazon bulk processor matches columns by HEADER NAME, not position — so appending the new
column at the end (col 28) is safe.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl import Workbook

SOURCE = 'bulk_upload_update_final_FIXED_v6.xlsx'
DEST   = 'bulk_upload_update_final_FIXED_v8.xlsx'
BROWSE_NODE = 'Mobile Displays (11180541031)'

print(f'Loading {SOURCE}...')
src_wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
src_ws = src_wb['Template']
src_rows = list(src_ws.iter_rows(values_only=True))
header = list(src_rows[0])
data_rows = src_rows[1:]
print(f'  {len(header)} cols, {len(data_rows)} SKUs')

# Append the new column
new_header = header + ['recommended_browse_nodes']
print(f'New header: {len(new_header)} cols (added "recommended_browse_nodes")')

# Build dest
dst_wb = Workbook()
dst_ws = dst_wb.active
dst_ws.title = 'Template'
dst_ws.append(new_header)

for r in data_rows:
    dst_ws.append(list(r) + [BROWSE_NODE])

dst_wb.save(DEST)
print(f'\nWrote {DEST}')
print(f'  Cols: {len(new_header)} (28 from v6 + 1 browse_node)')
print(f'  SKUs: {len(data_rows)}, all set to {BROWSE_NODE!r}')

# Verify
verify_wb = openpyxl.load_workbook(DEST, read_only=True, data_only=True)['Template']
last_col = None
for i, r in enumerate(verify_wb.iter_rows(values_only=True)):
    if i == 0:
        last_col = r[-1]
        print(f'\nVerify header col 28 = {r[-1]!r}')
    if i == 1:
        print(f'Verify SKU {r[0]} col 28 = {r[-1]!r}')
    if i == 1283:
        print(f'Verify last SKU {r[0]} col 28 = {r[-1]!r}')
        break
