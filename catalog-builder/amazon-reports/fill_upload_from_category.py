"""
Fill upload file - MINIMAL approach
Only writes columns that Amazon needs, skips full template copy
"""
import openpyxl

print("=" * 60)
print("FILL UPLOAD — MINIMAL FILE")
print("=" * 60)

COL = {
    'sku': 1, 'pt': 2, 'action': 3, 'name': 7, 'brand': 8,
    'img': 20, 'desc': 30, 'b1': 31, 'b2': 32, 'b3': 33, 'b4': 34, 'b5': 35, 'kw': 36,
    'c1': 44, 'c2': 45, 'c3': 46,
    'ful': 97, 'qty': 98, 'price': 102, 'mrp': 103
}

def calc_mrp(p):
    try:
        p = float(p)
        if p < 500:     return round(p * 5,    -2)
        elif p <= 2000: return round(round(p * 2, -2), -2)
        else:           return round(p * 1.43, -2)
    except:
        return 0

# Build Category Report lookup
print("\n[1] Category Report...")
cat = openpyxl.load_workbook('Category+Listings+Report_05-12-2026.xlsm', read_only=True, data_only=True)
cat_ws = cat.active
cat_data = {}
for row in range(7, cat_ws.max_row + 1):
    sku = cat_ws.cell(row, 3).value
    if not sku: continue
    sk = str(sku).strip().lower()
    cat_data[sk] = {
        'img': cat_ws.cell(row, 29).value or '',
        'qty': int(cat_ws.cell(row, 420).value) if cat_ws.cell(row, 420).value else 0,
        'price': float(cat_ws.cell(row, 424).value) if cat_ws.cell(row, 424).value else 0
    }
cat.close()
print(f"   {len(cat_data)} SKUs")

# Read bullets file read-only
print("\n[2] Bullets file...")
bl = openpyxl.load_workbook('PHONE_ACCESSORY_FILLED_bullets_v2.xlsx', read_only=True, data_only=True)
bl_ws = bl.active
rows = []
for row in range(7, bl_ws.max_row + 1):
    sku = bl_ws.cell(row, COL['sku']).value
    if not sku: continue
    sk = str(sku).strip().lower()
    cr = cat_data.get(sk)

    if cr:
        qty, price, img = cr['qty'], cr['price'], cr['img']
    else:
        qty = bl_ws.cell(row, COL['qty']).value or 0
        price = bl_ws.cell(row, COL['price']).value or 0
        try: price = float(price)
        except: price = 0
        img = ''

    rows.append({
        'sku': str(sku).strip(),
        'title': bl_ws.cell(row, COL['name']).value or '',
        'pt': bl_ws.cell(row, COL['pt']).value or 'PHONE_ACCESSORY',
        'action': bl_ws.cell(row, COL['action']).value or 'Edit (Partial Update)',
        'img': img,
        'qty': qty, 'price': price, 'mrp': calc_mrp(price),
        'b1': bl_ws.cell(row, COL['b1']).value or '',
        'b2': bl_ws.cell(row, COL['b2']).value or '',
        'b3': bl_ws.cell(row, COL['b3']).value or '',
        'b4': bl_ws.cell(row, COL['b4']).value or '',
        'b5': bl_ws.cell(row, COL['b5']).value or '',
        'kw': bl_ws.cell(row, COL['kw']).value or '',
        'c1': bl_ws.cell(row, COL['c1']).value or '',
        'c2': bl_ws.cell(row, COL['c2']).value or '',
        'c3': bl_ws.cell(row, COL['c3']).value or '',
    })
bl.close()
print(f"   {len(rows)} SKUs")

matched = sum(1 for r in rows if str(r['sku']).lower() in cat_data)
print(f"   {matched} matched with Category Report")

# Write minimal output
print("\n[3] Writing output...")
out = openpyxl.Workbook()
out_ws = out.active
out_ws.title = 'Template'

# Header row 4 (labels)
labels = ['SKU','Product Type','Listing Action','Parentage Level','Parent SKU',
          'Variation Theme Name','Item Name','Brand Name','Product Id Type',
          'Product Id'] + ['']*10 + ['Main Image URL'] + ['']*9 + \
          ['Product Description','Bullet Point','Bullet Point','Bullet Point',
           'Bullet Point','Bullet Point','Generic Keyword'] + ['']*7 + \
          ['Compatible Phone Models','Compatible Phone Models','Compatible Phone Models'] + \
          ['']*50 + ['Fulfillment Channel Code (IN)','Quantity (IN)'] + ['']*3 + \
          ['Your Price INR (Sell on Amazon, IN)','Maximum Retail Price (Sell on Amazon, IN)']
for c, lbl in enumerate(labels, 1):
    out_ws.cell(4, c).value = lbl

# Attribute row 5 (field names)
attrs = ['contribution_sku#1.value','product_type#1.value','::record_action',
         'parentage_level[marketplace_id=A21TJRUUN4KGV]#1.value'] + ['']*4 + \
         ['item_name[marketplace_id=A21TJRUUN4KGV][language_tag=en_IN]#1.value'] + \
         ['brand[marketplace_id=A21TJRUUN4KGV][language_tag=en_IN]#1.value'] + \
         ['']*11 + ['main_product_image_locator[marketplace_id=A21TJRUUN4KGV]#1.media_location'] + \
         ['']*9 + ['product_description[marketplace_id=A21TJRUUN4KGV][language_tag=en_IN]#1.value'] + \
         ['bullet_point[marketplace_id=A21TJRUUN4KGV][language_tag=en_IN]#1.value',
          'bullet_point[marketplace_id=A21TJRUUN4KGV][language_tag=en_IN]#2.value',
          'bullet_point[marketplace_id=A21TJRUUN4KGV][language_tag=en_IN]#3.value',
          'bullet_point[marketplace_id=A21TJRUUN4KGV][language_tag=en_IN]#4.value',
          'bullet_point[marketplace_id=A21TJRUUN4KGV][language_tag=en_IN]#5.value',
          'generic_keyword[marketplace_id=A21TJRUUN4KGV][language_tag=en_IN]#1.value'] + \
         ['']*7 + \
         ['compatible_phone_models[marketplace_id=A21TJRUUN4KGV][language_tag=en_IN]#1.value',
          'compatible_phone_models[marketplace_id=A21TJRUUN4KGV][language_tag=en_IN]#2.value',
          'compatible_phone_models[marketplace_id=A21TJRUUN4KGV][language_tag=en_IN]#3.value'] + \
         ['']*50 + \
         ['fulfillment_availability#1.fulfillment_channel_code',
          'fulfillment_availability#1.quantity'] + ['']*3 + \
         ['purchasable_offer[marketplace_id=A21TJRUUN4KGV][audience=ALL]#1.our_price#1.schedule#1.value_with_tax',
          'purchasable_offer[marketplace_id=A21TJRUUN4KGV][audience=ALL]#1.maximum_retail_price#1.schedule#1.value_with_tax']
for c, attr in enumerate(attrs, 1):
    out_ws.cell(5, c).value = attr

DESC = 'Compatible mobile phone display screen replacement. Tested for quality and performance before dispatch. Professional installation recommended for best results.'

# Data
for i, r in enumerate(rows):
    row = 7 + i
    out_ws.cell(row, COL['sku']).value = r['sku']
    out_ws.cell(row, COL['pt']).value = r['pt']
    out_ws.cell(row, COL['action']).value = r['action']
    out_ws.cell(row, COL['name']).value = r['title']
    # Brand = BLANK for Edit
    out_ws.cell(row, COL['img']).value = r['img']
    out_ws.cell(row, COL['desc']).value = DESC
    out_ws.cell(row, COL['b1']).value = r['b1']
    out_ws.cell(row, COL['b2']).value = r['b2']
    out_ws.cell(row, COL['b3']).value = r['b3']
    out_ws.cell(row, COL['b4']).value = r['b4']
    out_ws.cell(row, COL['b5']).value = r['b5']
    out_ws.cell(row, COL['kw']).value = r['kw']
    out_ws.cell(row, COL['c1']).value = r['c1']
    out_ws.cell(row, COL['c2']).value = r['c2']
    out_ws.cell(row, COL['c3']).value = r['c3']
    out_ws.cell(row, COL['ful']).value = 'Fulfillment by Merchant (Default)'
    out_ws.cell(row, COL['qty']).value = r['qty']
    out_ws.cell(row, COL['price']).value = r['price']
    out_ws.cell(row, COL['mrp']).value = r['mrp']

out_path = 'PHONE_ACCESSORY_upload.xlsx'
out.save(out_path)
out.close()
print(f"   Saved {out_path}")

# Quick verify
print("\n[4] Verify...")
vb = openpyxl.load_workbook(out_path, read_only=True, data_only=True)
vw = vb.active
count = 0
for row in range(7, vw.max_row + 1):
    sku = vw.cell(row, COL['sku']).value
    if not sku: continue
    qty = vw.cell(row, COL['qty']).value
    price = vw.cell(row, COL['price']).value
    mrp = vw.cell(row, COL['mrp']).value
    img = vw.cell(row, COL['img']).value
    b1 = (vw.cell(row, COL['b1']).value or '')[:70]
    c1 = vw.cell(row, COL['c1']).value or ''
    print(f"  {sku}: qty={qty}, Rs{price}->MRP Rs{mrp}, img={'YES' if img else 'NO'}, c1={c1[:30]}")
    print(f"    b1: {b1}...")
    count += 1
    if count >= 5: break
vb.close()

print(f"\n{'=' * 60}")
print("DONE!")
print(f"File: {out_path}")
print(f"SKUs: {len(rows)} ({matched} matched)")
print("Ready!")