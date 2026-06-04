"""
Amazon Bulk Listing Upload - Simplified Version
"""
import openpyxl
import re

print("=" * 70)
print("AMAZON BULK LISTING UPLOAD GENERATOR")
print("=" * 70)

# Config
CONFIG = {
    'product_type': 'PHONE_ACCESSORY',
    'listing_action': 'Edit (Partial Update)',
    'brand_name': '',  # BLANK for Edit
    'fulfillment_channel': 'Fulfillment by Merchant (Default)',
    'description': 'Compatible mobile phone display screen replacement. Tested for quality and performance before dispatch. Professional installation recommended for best results.',
}

def calculate_mrp(price):
    if price < 500:
        return round(price * 5, -2)
    elif price <= 2000:
        return round(price * 2, -2)
    else:
        return round(price * 1.43, -2)

def detect_tier(title):
    t = str(title).lower()
    if 'careog' in t:
        return 'gold'
    elif 'oled' in t or 'amoled' in t:
        return 'silver'
    return 'bronze'

def detect_frame(title):
    t = str(title).lower()
    return 'with frame' in t or 'folder' in t

def detect_fingerprint(title):
    t = str(title).lower()
    return ('fingerprint support' in t or 'oled' in t) and 'no fingerprint' not in t

def extract_model(title):
    match = re.search(r'Compatible for\s+(.+?)(?:\s+(?:LCD|OLED|Display|AMOLED|CareOG)|$)', str(title), re.IGNORECASE)
    if match:
        return re.sub(r'\s*\([^)]*\)\s*', ' ', match.group(1)).strip()
    return ''

def make_bullets(title, tier, frame, fingerprint, model):
    fp = ' — Fingerprint supported after fitting' if fingerprint else ''
    b1 = f"Compatible with {model}{fp} — Verify model number from Settings before ordering"

    if tier == 'gold':
        b2 = "Gold Tier CareOG — Tested OG combo with premium quality. Superior reliability for professional repair technicians."
    elif tier == 'silver':
        b2 = "Silver Tier OLED — Premium display with fingerprint support. Vivid colors and better contrast." if fingerprint else "Silver Tier OLED — Premium display quality with vivid colors and better contrast."
    else:
        b2 = "Bronze Tier LCD — Budget friendly good quality display. Great value for money, reliable performance."

    frame_note = " With frame for easy installation." if frame else ""
    if tier == 'gold':
        b3 = f"Quality Tested — 100% checked for dead pixels, color calibration, touch response.{frame_note}"
    elif tier == 'silver':
        b3 = f"OLED Quality — Original chip IC, vibrant colors, tested panel.{frame_note}"
    else:
        b3 = f"LCD Quality — Good color reproduction, original chip compatible, tested.{frame_note}"

    if frame:
        b4 = "With Frame Assembly — Screen pre-pasted on frame, ready to install. Original adhesive pre-applied." if not fingerprint else "With Frame Assembly — Pre-framed screen for fingerprint calibration. Ready to install."
    else:
        b4 = "Screen Only — Requires frame transfer from old screen. Adhesive sheet included. Professional installation recommended."

    b5 = "Warranty: 7 days replacement — QC tested before dispatch. No warranty after film removal. For help visit Repairfully.com"

    return [b1, b2, b3, b4, b5]

def make_keywords(tier, fingerprint):
    base = "display sceen replacement mobile repair"
    tier_kw = "careog tested combo premium quality" if tier == 'gold' else "oled amoled premium fingerprint" if tier == 'silver' else "lcd budget friendly non fingerprint"
    fp_kw = "fingerprint support" if fingerprint else "non fingerprint"
    return f"{base} {tier_kw} {fp_kw}"

# Read Category Report
print("\n[1] Reading Category Report...")
wb = openpyxl.load_workbook('Category+Listings+Report_05-12-2026.xlsm', read_only=True, data_only=True)
ws = wb.active

listings = []
for row in range(7, ws.max_row + 1):
    sku = ws.cell(row, 3).value
    title = ws.cell(row, 2).value
    if not sku or not title:
        continue

    price = float(ws.cell(row, 424).value) if ws.cell(row, 424).value else 0
    qty = int(ws.cell(row, 420).value) if ws.cell(row, 420).value else 0
    img = ws.cell(row, 29).value or ''

    try:
        price = float(price)
    except:
        price = 0

    tier = detect_tier(title)
    frame = detect_frame(title)
    fp = detect_fingerprint(title)
    model = extract_model(title)
    mrp = calculate_mrp(price)
    bullets = make_bullets(title, tier, frame, fp, model)
    keywords = make_keywords(tier, fp)

    listings.append({
        'sku': str(sku).strip(),
        'title': str(title).strip(),
        'price': float(price),
        'mrp': float(mrp),
        'qty': int(qty),
        'img': img,
        'tier': tier,
        'frame': frame,
        'fp': fp,
        'model': model,
        'bullets': bullets,
        'keywords': keywords
    })

wb.close()
print(f"   Read {len(listings)} SKUs")

# Create output
print("\n[2] Creating bulk upload file...")

# Copy template
template = openpyxl.load_workbook('PHONE_ACCESSORY.xlsm')
template_ws = template['Template']

out = openpyxl.Workbook()
out_ws = out.active
out_ws.title = 'Template'

# Copy rows 1-6
for r in range(1, 7):
    for c in range(1, 235):
        out_ws.cell(r, c).value = template_ws.cell(r, c).value

# Write data
COL = {'sku':1, 'pt':2, 'action':3, 'name':7, 'brand':8, 'img':20, 'desc':30,
       'b1':31, 'b2':32, 'b3':33, 'b4':34, 'b5':35, 'kw':36,
       'c1':44, 'c2':45, 'c3':46, 'ful':419, 'qty':420, 'price':424, 'mrp':425}

for i, l in enumerate(listings):
    r = 7 + i
    out_ws.cell(r, COL['sku']).value = l['sku']
    out_ws.cell(r, COL['pt']).value = CONFIG['product_type']
    out_ws.cell(r, COL['action']).value = CONFIG['listing_action']
    out_ws.cell(r, COL['name']).value = l['title']
    # Brand = BLANK for Edit
    out_ws.cell(r, COL['img']).value = l['img']
    out_ws.cell(r, COL['desc']).value = CONFIG['description']
    out_ws.cell(r, COL['b1']).value = l['bullets'][0]
    out_ws.cell(r, COL['b2']).value = l['bullets'][1]
    out_ws.cell(r, COL['b3']).value = l['bullets'][2]
    out_ws.cell(r, COL['b4']).value = l['bullets'][3]
    out_ws.cell(r, COL['b5']).value = l['bullets'][4]
    out_ws.cell(r, COL['kw']).value = l['keywords']
    out_ws.cell(r, COL['c1']).value = l['model']
    # c2, c3 = cross-fits (empty for now)
    out_ws.cell(r, COL['ful']).value = CONFIG['fulfillment_channel']
    out_ws.cell(r, COL['qty']).value = l['qty']
    out_ws.cell(r, COL['price']).value = l['price']
    out_ws.cell(r, COL['mrp']).value = l['mrp']

out.save('bulk_upload_update.xlsx')
out.close()
template.close()

# Summary
tier_counts = {'gold': 0, 'silver': 0, 'bronze': 0}
for l in listings:
    tier_counts[l['tier']] += 1

print(f"\n{'=' * 70}")
print("DONE!")
print(f"{'=' * 70}")
print(f"File: bulk_upload_update.xlsx")
print(f"Total SKUs: {len(listings)}")
print(f"Gold (CareOG): {tier_counts['gold']}")
print(f"Silver (OLED): {tier_counts['silver']}")
print(f"Bronze (LCD): {tier_counts['bronze']}")

# Sample
print(f"\nSample (first 3):")
for l in listings[:3]:
    print(f"  {l['sku']}: ₹{l['price']:.0f} → MRP ₹{l['mrp']:.0f}, {l['tier'].upper()}, {l['model'][:30]}")

print("\nReady for upload!")