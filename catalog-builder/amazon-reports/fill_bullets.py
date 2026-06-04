import openpyxl
import re
from datetime import datetime

# Load workbook
print("Loading PHONE_ACCESSORY_FILLED.xlsm...")
wb = openpyxl.load_workbook('PHONE_ACCESSORY_FILLED.xlsm', data_only=True)
ws = wb['Template']

# Column indices (1-based)
COL_SKU = 1
COL_ITEM_NAME = 7
COL_BULLET1 = 31
COL_BULLET2 = 32
COL_BULLET3 = 33
COL_BULLET4 = 34
COL_BULLET5 = 35
COL_KEYWORD = 36

def extract_model_info(title):
    """Extract device model from item name title."""
    if not title:
        return None, None, None

    title_str = str(title)

    # Detect screen tech
    if 'OLED' in title_str:
        tech = 'OLED'
        fingerprint = 'Fingerprint Supported'
    else:
        tech = 'LCD'
        fingerprint = 'Non Fingerprint'

    # Detect frame variant
    if 'Folder' in title_str or 'WF' in title_str or 'With Frame' in title_str:
        frame = 'With Frame'
    else:
        frame = 'Without Frame'

    # Extract brand and model
    # Pattern: "Compatible for [Brand] [Model] ..." or "Compatible for [Brand] [Model] (variant) ..."
    # Stop at first occurrence of LCD/OLED/Display
    match = re.search(r'Compatible for\s+([A-Za-z0-9\-\s]+?)(?:\s+(?:LCD|OLED)|$)', title_str)
    if match:
        device_str = match.group(1).strip()
        # Remove parenthetical content like "(No Fingerprint Support)"
        device_str = re.sub(r'\s*\([^)]*\)\s*', ' ', device_str).strip()
        # Split into brand and model (first word = brand, rest = model)
        parts = device_str.split(' ', 1)
        if len(parts) >= 2:
            brand = parts[0]
            model = parts[1].strip()
        else:
            brand = parts[0]
            model = parts[0]
    else:
        # Fallback: just get everything after "Compatible for" until LCD/OLED/Display
        fallback = re.sub(r'Compatible for\s+', '', title_str)
        fallback = re.split(r'\s+(?:LCD|OLED|Display)', fallback)[0].strip()
        fallback = re.sub(r'\s*\([^)]*\)\s*', ' ', fallback).strip()
        parts = fallback.split(' ', 1)
        brand = parts[0] if parts else None
        model = parts[1] if len(parts) > 1 else fallback

    return brand, model, (tech, fingerprint, frame)

def generate_bullet1(brand, model, tech):
    """Bullet 1: Compatibility."""
    if not model:
        return ""

    model_clean = re.sub(r'[()\-\s]+', ' ', model).strip()

    if brand and brand.lower() in ['iphone', 'apple']:
        return f"Compatible with {brand} {model_clean} — Check Settings > General > About for exact Model Number before ordering"
    elif brand and brand.lower() in ['samsung']:
        return f"Compatible with Samsung {model_clean} — Check Settings > About Phone for exact Model Number before ordering"
    else:
        return f"Compatible with {brand} {model_clean} — Verify your model number from Settings before ordering"

def generate_bullet2(brand, model, tech, fingerprint):
    """Bullet 2: Screen Tech + Fingerprint."""
    device_str = f"{brand} {model}" if model else ""

    if fingerprint == 'Non Fingerprint':
        return f"{device_str} {tech} Display — Non Fingerprint. {tech} screens have no under-display fingerprint sensor. Standard installation, no fingerprint calibration needed."
    else:
        return f"{device_str} {tech} Display — Fingerprint Supported. {tech} screens support under-display fingerprint but require framing during fitting. Fingerprint takes 1-2 days to adjust after installation and calibration."

def generate_bullet3():
    """Bullet 3: Quality Grade."""
    return "CareOG Quality — Premium tested display with original chip IC, 100% checked for dead pixels, colour calibration, and touch response before dispatch."

def generate_bullet4(frame, fingerprint):
    """Bullet 4: Build Variant + Installation."""
    if frame == 'With Frame':
        if fingerprint == 'Fingerprint Supported':
            return "With Frame (WF) — Pre-framed assembly for fingerprint calibration. Frame transfer not needed. Adhesive included. Professional install recommended for fingerprint setup."
        else:
            return "With Frame (WF) — Screen pre-pasted on frame, ready to install. Original adhesive pre-applied. Colored frame rims shipped based on availability."
    else:
        return "Without Frame — Screen only. Requires frame transfer from your old screen. Adhesive sheet included. Professional installation recommended."

def generate_bullet5():
    """Bullet 5: Warranty + GST + Bulk."""
    return "Warranty: 30 days replacement — GST invoice available on request — Bulk order pricing on WhatsApp — Order before 2 PM for same day dispatch"

def generate_keywords(brand, tech):
    """Generate backend keywords (200 bytes max)."""
    if brand and brand.lower() in ['iphone', 'apple']:
        base = "apple iphone display sceen replacement a-code oled lcd fingerprint folder marammat mobile repair"
    elif brand and brand.lower() in ['samsung']:
        base = "samsung galaxy display sceen replacement sm-code oled lcd fingerprint folder marammat mobile repair"
    else:
        if tech == 'OLED':
            base = "display sceen replacement oled fingerprint marammat mobile repair folder"
        else:
            base = "display sceen replacement lcd non fingerprint marammat mobile repair folder"

    return base

# Process all rows
print("\nProcessing listings...")
filled_count = 0
skipped_count = 0
errors = []

for row in range(7, ws.max_row + 1):
    sku = ws.cell(row, COL_SKU).value
    title = ws.cell(row, COL_ITEM_NAME).value

    if not title or not str(title).strip():
        skipped_count += 1
        continue

    try:
        # Extract model info
        brand, model, (tech, fingerprint, frame) = extract_model_info(title)

        # Generate bullets
        bullet1 = generate_bullet1(brand, model, tech)
        bullet2 = generate_bullet2(brand, model, tech, fingerprint)
        bullet3 = generate_bullet3()
        bullet4 = generate_bullet4(frame, fingerprint)
        bullet5 = generate_bullet5()

        # Generate keywords
        keywords = generate_keywords(brand, tech)

        # Write to cells
        ws.cell(row, COL_BULLET1).value = bullet1
        ws.cell(row, COL_BULLET2).value = bullet2
        ws.cell(row, COL_BULLET3).value = bullet3
        ws.cell(row, COL_BULLET4).value = bullet4
        ws.cell(row, COL_BULLET5).value = bullet5
        ws.cell(row, COL_KEYWORD).value = keywords

        filled_count += 1

    except Exception as e:
        errors.append(f"Row {row} ({sku}): {str(e)}")
        continue

print(f"\n{'='*60}")
print("BULLET POINTS FILL COMPLETE")
print(f"{'='*60}")
print(f"Filled: {filled_count} listings")
print(f"Skipped (empty): {skipped_count}")
if errors:
    print(f"Errors: {len(errors)}")
    for err in errors[:5]:
        print(f"  {err}")

# Save
output_file = 'PHONE_ACCESSORY_FILLED_bullets.xlsm'
print(f"\nSaving to {output_file}...")
wb.save(output_file)
wb.close()
print("Done!")

# Print sample
print(f"\n{'='*60}")
print("SAMPLE OUTPUT (first listing):")
print(f"{'='*60}")
wb2 = openpyxl.load_workbook(output_file, data_only=True)
ws2 = wb2['Template']
sample_title = ws2.cell(7, COL_ITEM_NAME).value
sample_b1 = ws2.cell(7, COL_BULLET1).value
sample_b2 = ws2.cell(7, COL_BULLET2).value
sample_b3 = ws2.cell(7, COL_BULLET3).value
sample_b4 = ws2.cell(7, COL_BULLET4).value
sample_b5 = ws2.cell(7, COL_BULLET5).value
sample_kw = ws2.cell(7, COL_KEYWORD).value

print(f"Title: {sample_title}")
print(f"\nBullet 1:\n  {sample_b1}")
print(f"\nBullet 2:\n  {sample_b2}")
print(f"\nBullet 3:\n  {sample_b3}")
print(f"\nBullet 4:\n  {sample_b4}")
print(f"\nBullet 5:\n  {sample_b5}")
print(f"\nKeywords:\n  {sample_kw}")

# Character counts
print(f"\n{'='*60}")
print("CHARACTER COUNTS:")
print(f"{'='*60}")
for i, (b, text) in enumerate([(1, sample_b1), (2, sample_b2), (3, sample_b3), (4, sample_b4), (5, sample_b5)], 1):
    if text:
        print(f"  Bullet {i}: {len(text)} chars {'✓' if len(text) <= 255 else '✗ OVER LIMIT'}")

wb2.close()