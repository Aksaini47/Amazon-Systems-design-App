"""Fixed bulk upload - minimal version"""
import openpyxl, csv, re

print("BULK LISTING UPLOAD - FIXED")
print("=" * 50)

# Load cross-fits
crossfits = {}
for fname in ['stock_1_exact_match.csv', 'stock_2_loose_match.csv', 'stock_3_not_listed.csv', 'stock_A_combined.csv']:
    try:
        with open(fname, 'r') as f:
            for r in csv.DictReader(f):
                cf = r.get('Sunsky Cross-fit Models', '').strip()
                sku = r.get('Amazon SKU', '').strip()
                if cf and sku:
                    crossfits[sku] = [c.strip() for c in cf.split('/') if c.strip()]
    except: pass
print(f"Cross-fits: {len(crossfits)}")

# Load Category Report
print("Loading Category Report...")
cat_wb = openpyxl.load_workbook('Category+Listings+Report_05-12-2026.xlsm', read_only=True)
cat_ws = cat_wb.active
print(f"Rows: {cat_ws.max_row}, Cols: {cat_ws.max_column}")

# Read needed columns: 2=title, 3=sku, 29-38=images, 420=qty, 424=price
print("Reading data rows...")
cat_data = {}

for row_data in cat_ws.iter_rows(min_row=7, max_row=cat_ws.max_row,
                                  min_col=2, max_col=424,
                                  values_only=True):
    # row_data[0]=col2, row_data[1]=col3, ..., row_data[422]=col424
    sku = row_data[1]  # col 3
    if not sku: continue

    # Images at cols 29-38 → indices 27-36
    imgs = [row_data[i] for i in range(27, 37) if i < len(row_data) and row_data[i]]
    # qty at col 420 → index 418
    qty_val = row_data[418] if 418 < len(row_data) else 0
    # price at col 424 → index 422
    price_val = row_data[422] if 422 < len(row_data) else 0

    cat_data[str(sku).strip().lower()] = {
        'sku': str(sku).strip(),
        'title': str(row_data[0] or '').strip(),
        'price': float(price_val) if price_val else 0,
        'qty': int(qty_val) if qty_val else 0,
        'images': [str(i).strip() for i in imgs if i]
    }

cat_wb.close()
print(f"Loaded {len(cat_data)} SKUs")
print(f"Done loading!")

# Generate listings
def calc_mrp(p):
    p = float(p)
    if p < 500: return round(p * 5, -2)
    elif p <= 2000: return round(p * 2, -2)
    else: return round(p * 1.43, -2)

def get_model(t):
    m = re.search(r'Compatible for\s+(.+?)(?:\s+(?:LCD|OLED|Display|AMOLED|CareOG|Super)|$)', str(t), re.IGNORECASE)
    if m: return re.sub(r'\s*\([^)]*\)\s*', ' ', m.group(1)).strip()
    return ''

def has_frame(t):
    t = str(t).lower()
    return '(with frame)' in t or ' with frame' in t

def has_fp(t):
    t = str(t).lower()
    return ('oled' in t or 'amoled' in t or 'fingerprint support' in t) and 'no fingerprint' not in t

print("Generating...")
listings = []
for sku_key, data in cat_data.items():
    sku = data['sku']
    title = data['title']
    model = get_model(title)
    frame = has_frame(title)
    fp = has_fp(title)
    imgs = data['images']

    fp_txt = ' — Fingerprint supported after fitting' if fp else ' — Fingerprint not supported'
    b1 = f"Compatible with {model}{fp_txt} — Verify model number from Settings before ordering"

    if frame:
        b2 = "Premium quality combo — LCD display with touch panel, original chip IC. Pre-pasted assembly for easy installation."
        b3 = "QC tested — Dead pixel check, color calibration, touch response verified. With frame for fingerprint calibration."
        b4 = "With Frame Assembly — Screen pre-pasted on frame, ready to install. Original adhesive applied."
    else:
        b2 = ""
        b3 = ""
        b4 = ""

    b5 = "Warranty: 7 days replacement — QC tested before dispatch. No warranty after film removal or installation. For help visit Repairfully.com"

    kw_base = "display sceen replacement mobile repair"
    kw_dtype = "careog tested combo" if 'careog' in str(title).lower() else "oled amoled" if 'oled' in str(title).lower() or 'amoled' in str(title).lower() else "lcd"
    kw_fp = "fingerprint support" if fp else "non fingerprint"
    kw_model = model.lower()
    kw = f"{kw_base} {kw_dtype} {kw_fp} {kw_model}"[:200]

    cf = crossfits.get(sku, [])
    c1 = model
    c2 = cf[0] if len(cf) > 0 else ''
    c3 = cf[1] if len(cf) > 1 else ''

    listings.append({
        'sku': sku, 'title': title, 'price': data['price'],
        'mrp': calc_mrp(data['price']), 'qty': data['qty'],
        'img_main': imgs[0] if len(imgs) > 0 else '',
        'img2': imgs[1] if len(imgs) > 1 else '',
        'img3': imgs[2] if len(imgs) > 2 else '',
        'img4': imgs[3] if len(imgs) > 3 else '',
        'img5': imgs[4] if len(imgs) > 4 else '',
        'img6': imgs[5] if len(imgs) > 5 else '',
        'img7': imgs[6] if len(imgs) > 6 else '',
        'img8': imgs[7] if len(imgs) > 7 else '',
        'b1': b1, 'b2': b2, 'b3': b3, 'b4': b4, 'b5': b5,
        'kw': kw, 'model': model, 'c1': c1, 'c2': c2, 'c3': c3
    })
    if len(listings) % 200 == 0: print(f"  {len(listings)}...")

print(f"Generated {len(listings)} listings")
print("Done generating!")

# Write output
print("Writing output...")
out = openpyxl.Workbook()
ws = out.active
ws.title = 'Template'

headers = ['SKU','product_type','listing_action','item_name','brand',
           'main_image_url','other_image_url_1','other_image_url_2',
           'other_image_url_3','other_image_url_4','other_image_url_5',
           'other_image_url_6','other_image_url_7','other_image_url_8',
           'description','bullet1','bullet2','bullet3','bullet4','bullet5',
           'keywords','compatible_phone_1','compatible_phone_2','compatible_phone_3',
           'fulfillment_channel','quantity','your_price','mrp']

for c, h in enumerate(headers, 1):
    ws.cell(1, c).value = h

DESC = "Compatible mobile phone display screen replacement. Tested for quality and performance before dispatch. Professional installation recommended."

for i, l in enumerate(listings):
    row = i + 2
    ws.cell(row, 1).value = l['sku']
    ws.cell(row, 2).value = 'PHONE_ACCESSORY'
    ws.cell(row, 3).value = 'Edit (Partial Update)'
    ws.cell(row, 4).value = l['title']
    ws.cell(row, 5).value = ''
    ws.cell(row, 6).value = l['img_main']
    ws.cell(row, 7).value = l['img2']
    ws.cell(row, 8).value = l['img3']
    ws.cell(row, 9).value = l['img4']
    ws.cell(row, 10).value = l['img5']
    ws.cell(row, 11).value = l['img6']
    ws.cell(row, 12).value = l['img7']
    ws.cell(row, 13).value = l['img8']
    ws.cell(row, 14).value = DESC
    ws.cell(row, 15).value = l['b1']
    ws.cell(row, 16).value = l['b2']
    ws.cell(row, 17).value = l['b3']
    ws.cell(row, 18).value = l['b4']
    ws.cell(row, 19).value = l['b5']
    ws.cell(row, 20).value = l['kw']
    ws.cell(row, 21).value = l['c1']
    ws.cell(row, 22).value = l['c2']
    ws.cell(row, 23).value = l['c3']
    ws.cell(row, 24).value = 'Fulfillment by Merchant (Default)'
    ws.cell(row, 25).value = l['qty']
    ws.cell(row, 26).value = l['price']
    ws.cell(row, 27).value = l['mrp']
    if i % 200 == 0: print(f"  Row {row}...")

out.save('bulk_upload_update_v2.xlsx')
out.close()

print(f"SAVED! {len(listings)} listings, 27 columns")
print("DONE!")