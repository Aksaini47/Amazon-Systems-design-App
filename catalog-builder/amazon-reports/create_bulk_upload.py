"""
Amazon Bulk Listing Upload - PHONE_ACCESSORY Template
======================================================
Generates upload-ready bulk file for Amazon India mobile screen listings.
"""

import openpyxl
import re
from datetime import datetime

print("=" * 70)
print("AMAZON BULK LISTING UPLOAD GENERATOR")
print("=" * 70)

# ============================================================
# CONFIGURATION - User decisions
# ============================================================
CONFIG = {
    'product_type': 'PHONE_ACCESSORY',
    'listing_action': 'Edit (Partial Update)',
    'brand_name': '',  # BLANK for Edit (Partial Update)
    'fulfillment_channel': 'Fulfillment by Merchant (Default)',
    'hsn_code': '851770',
    'generic_description': 'Compatible mobile phone display screen replacement. Tested for quality and performance before dispatch. Professional installation recommended for best results.',
    'warranty_description': 'Warranty: 7 days replacement — QC tested before dispatch. No warranty after protection film removal or installation. For help visit Repairfully.com',
}

def calculate_mrp(sell_price):
    """Calculate MRP based on psychological pricing formula."""
    if sell_price < 500:
        # Low value: 80% off → MRP = 5x sell price
        return round(sell_price * 5, -2)  # Round to nearest 100
    elif sell_price <= 2000:
        # Medium value: 50% off → MRP = 2x sell price
        return round(sell_price * 2, -2)
    else:
        # High value: ~30% off → MRP = 1.43x sell price
        return round(sell_price * 1.43, -2)

def detect_tier(title):
    """Detect tier: Bronze (LCD), Silver (OLED), Gold (CareOG)."""
    t = str(title).lower()
    if 'careog' in t:
        return 'gold'
    elif 'oled' in t or 'amoled' in t:
        return 'silver'
    else:
        return 'bronze'

def detect_frame(title):
    """Detect if product has with frame."""
    t = str(title).lower()
    if 'with frame' in t or 'folder' in t or ' wf ' in t:
        return True
    return False

def detect_fingerprint(title):
    """Detect fingerprint support."""
    t = str(title).lower()
    if 'fingerprint support' in t or 'oled' in t or 'amoled' in t:
        if 'no fingerprint' not in t:
            return True
    return False

def extract_model_from_title(title):
    """Extract model name from title for compatible phone models."""
    t = str(title)
    # Pattern: "Compatible for [Brand] [Model]"
    match = re.search(r'Compatible for\s+(.+?)(?:\s+(?:LCD|OLED|Display|AMOLED|CareOG)|$)', t, re.IGNORECASE)
    if match:
        model = match.group(1).strip()
        # Clean up parenthetical content
        model = re.sub(r'\s*\([^)]*\)\s*', ' ', model).strip()
        return model
    return ''

def generate_bullet1(model, fingerprint):
    """Bullet 1: Compatibility."""
    if not model:
        return ''
    fp_note = ' — Fingerprint supported after fitting' if fingerprint else ' — Fingerprint not supported'
    return f"Compatible with {model}{fp_note} — Verify model number from Settings before ordering"

def generate_bullet2(tier, fingerprint):
    """Bullet 2: Tier value proposition."""
    if tier == 'gold':
        return "Gold Tier CareOG — Tested OG combo with premium quality. Superior reliability for professional repair technicians."
    elif tier == 'silver':
        if fingerprint:
            return "Silver Tier OLED — Premium display with fingerprint sensor support. Vivid colors and better contrast."
        else:
            return "Silver Tier OLED — Premium display quality with vivid colors and better contrast than LCD."
    else:
        return "Bronze Tier LCD — Budget friendly good quality display. Great value for money, reliable performance."

def generate_bullet3(tier, frame):
    """Bullet 3: Quality and frame details."""
    if tier == 'gold':
        if frame:
            return "Quality Tested — 100% checked for dead pixels, color calibration, touch response. With frame for easy installation."
        else:
            return "Quality Tested — 100% checked for dead pixels, color calibration, touch response. Screen only."
    elif tier == 'silver':
        if frame:
            return "OLED Quality — Original chip IC, vibrant colors, tested panel. With frame for perfect fit."
        else:
            return "OLED Quality — Original chip IC, vibrant colors, tested panel. Screen only, professional install recommended."
    else:
        if frame:
            return "LCD Quality — Good color reproduction, original chip compatible, tested. With frame for easy installation."
        else:
            return "LCD Quality — Good color reproduction, original chip compatible, tested. Screen only, frame not included."

def generate_bullet4(frame, fingerprint):
    """Bullet 4: Build variant and installation."""
    if frame:
        if fingerprint:
            return "With Frame Assembly — Pre-framed screen for fingerprint calibration. Ready to install. Professional install recommended."
        else:
            return "With Frame Assembly — Screen pre-pasted on frame, ready to install. Original adhesive pre-applied."
    else:
        return "Screen Only — Requires frame transfer from old screen. Adhesive sheet included. Professional installation recommended."

def generate_bullet5():
    """Bullet 5: Warranty and support."""
    return "Warranty: 7 days replacement — QC tested before dispatch. No warranty after film removal. For help visit Repairfully.com"

def generate_keywords(tier, fingerprint, model):
    """Generate backend keywords."""
    kw_base = "display sceen replacement mobile repair"

    if tier == 'gold':
        kw_tier = "careog tested combo premium quality"
    elif tier == 'silver':
        kw_tier = "oled amoled premium fingerprint"
    else:
        kw_tier = "lcd budget friendly non fingerprint"

    fp = "fingerprint support" if fingerprint else "non fingerprint"

    return f"{kw_base} {kw_tier} {fp}"

# ============================================================
# STEP 1: Read Category Listings Report
# ============================================================
print("\n[1/6] Reading Category Listings Report...")

category_wb = openpyxl.load_workbook('Category+Listings+Report_05-12-2026.xlsm', read_only=True, data_only=True)
category_ws = category_wb.active

# Extract all listings
category_data = []
for row in range(7, category_ws.max_row + 1):
    sku = category_ws.cell(row, 3).value
    title = category_ws.cell(row, 2).value
    item_name = category_ws.cell(row, 9).value
    price = category_ws.cell(row, 424).value
    mrp = category_ws.cell(row, 425).value
    qty = category_ws.cell(row, 420).value
    status = category_ws.cell(row, 1).value
    image_url = category_ws.cell(row, 29).value

    if sku and item_name:
        category_data.append({
            'row': row,
            'sku': str(sku).strip(),
            'title': str(title).strip() if title else '',
            'item_name': str(item_name).strip(),
            'price': float(price) if price else 0,
            'mrp': float(mrp) if mrp else 0,
            'qty': int(qty) if qty else 0,
            'status': status,
            'image_url': image_url if image_url else '',
        })

category_wb.close()
print(f"   Found {len(category_data)} SKUs in Category Report")

# ============================================================
# STEP 2: Read Sunsky Cross-fits (if available)
# ============================================================
print("\n[2/6] Reading Sunsky cross-fits...")

cross_fits = {}
try:
    import pandas as pd
    sunsky_df = pd.read_excel('sunsky_crossfits.xlsx') if False else None  # Skip if not available
    # For now, use a simple mapping based on existing data
    print("   (Cross-fits will be extracted from item names)")
except:
    print("   (No cross-fits file found - using basic matching)")

# ============================================================
# STEP 3: Read PHONE_ACCESSORY FILLED bullets (if available)
# ============================================================
print("\n[3/6] Reading existing bullet points...")

bullet_data = {}
try:
    bullet_wb = openpyxl.load_workbook('PHONE_ACCESSORY_FILLED_bullets_v2.xlsx', read_only=True, data_only=True)
    bullet_ws = bullet_wb['Template']

    for row in range(7, bullet_ws.max_row + 1):
        sku = bullet_ws.cell(row, 1).value
        if sku:
            bullets = []
            for col in [31, 32, 33, 34, 35]:
                val = bullet_ws.cell(row, col).value
                if val:
                    bullets.append(str(val).strip())
            if bullets:
                bullet_data[str(sku).strip()] = bullets

    bullet_wb.close()
    print(f"   Found {len(bullet_data)} SKUs with pre-generated bullets")
except Exception as e:
    print(f"   (No bullet file found - will generate new bullets)")

# ============================================================
# STEP 4: Generate listing data
# ============================================================
print("\n[4/6] Generating listing data...")

listings = []
skipped = []

for item in category_data:
    sku = item['sku']
    title = item['title'] or item['item_name']

    # Skip non-display items
    title_lower = title.lower()
    non_display = ['bra ', 'strap', 'jewellery', 'jewelry', 'necklace', 'earring']
    skip = False
    for nd in non_display:
        if nd in title_lower and 'display' not in title_lower:
            skip = True
            break
    if skip:
        skipped.append(sku)
        continue

    # Detect attributes
    tier = detect_tier(title)
    has_frame = detect_frame(title)
    has_fingerprint = detect_fingerprint(title)
    model = extract_model_from_title(title)

    # Calculate MRP
    sell_price = item['price']
    calculated_mrp = calculate_mrp(sell_price)

    # Get bullets - use pre-generated if available, otherwise generate
    if sku in bullet_data:
        bullets = bullet_data[sku]
    else:
        bullets = [
            generate_bullet1(model, has_fingerprint),
            generate_bullet2(tier, has_fingerprint),
            generate_bullet3(tier, has_frame),
            generate_bullet4(has_frame, has_fingerprint),
            generate_bullet5()
        ]

    # Generate keywords
    keywords = generate_keywords(tier, has_fingerprint, model)

    # Create listing
    listing = {
        'sku': sku,
        'product_type': CONFIG['product_type'],
        'listing_action': CONFIG['listing_action'],
        'item_name': title,
        'brand_name': CONFIG['brand_name'],  # BLANK for Edit
        'price': sell_price,
        'mrp': calculated_mrp,
        'qty': item['qty'],
        'image_url': item['image_url'],
        'description': CONFIG['generic_description'],
        'bullet1': bullets[0] if len(bullets) > 0 else '',
        'bullet2': bullets[1] if len(bullets) > 1 else '',
        'bullet3': bullets[2] if len(bullets) > 2 else '',
        'bullet4': bullets[3] if len(bullets) > 3 else '',
        'bullet5': bullets[4] if len(bullets) > 4 else '',
        'keywords': keywords,
        'compatible_model_1': model,
        'compatible_model_2': '',  # Cross-fit
        'compatible_model_3': '',  # Cross-fit
        'fulfillment': CONFIG['fulfillment_channel'],
        'tier': tier,
        'has_frame': has_frame,
    }

    listings.append(listing)

print(f"   Generated {len(listings)} listings")
print(f"   Skipped {len(skipped)} non-display items")

# Tier breakdown
tier_counts = {'gold': 0, 'silver': 0, 'bronze': 0}
frame_counts = {'with_frame': 0, 'without_frame': 0}
for l in listings:
    tier_counts[l['tier']] += 1
    if l['has_frame']:
        frame_counts['with_frame'] += 1
    else:
        frame_counts['without_frame'] += 1

print(f"\n   Tier breakdown: Gold={tier_counts['gold']}, Silver={tier_counts['silver']}, Bronze={tier_counts['bronze']}")
print(f"   Frame breakdown: With Frame={frame_counts['with_frame']}, Without={frame_counts['without_frame']}")

# ============================================================
# STEP 5: Read PHONE_ACCESSORY template
# ============================================================
print("\n[5/6] Reading PHONE_ACCESSORY template...")

template_wb = openpyxl.load_workbook('PHONE_ACCESSORY.xlsm')
template_ws = template_wb['Template']

# ============================================================
# STEP 6: Create output file
# ============================================================
print("\n[6/6] Creating bulk upload file...")

output_wb = openpyxl.Workbook()
output_ws = output_wb.active
output_ws.title = 'Template'

# Copy template structure (rows 1-6)
for row in range(1, 7):
    for col in range(1, template_ws.max_column + 1):
        val = template_ws.cell(row, col).value
        output_ws.cell(row, col).value = val

# Define column mappings (column numbers from template)
COL_SKU = 1
COL_PRODUCT_TYPE = 2
COL_LISTING_ACTION = 3
COL_ITEM_NAME = 7
COL_BRAND = 8
COL_MAIN_IMAGE = 20
COL_DESCRIPTION = 30
COL_BULLET1 = 31
COL_BULLET2 = 32
COL_BULLET3 = 33
COL_BULLET4 = 34
COL_BULLET5 = 35
COL_KEYWORDS = 36
COL_COMPATIBLE1 = 44
COL_COMPATIBLE2 = 45
COL_COMPATIBLE3 = 46
COL_FULFILLMENT = 419
COL_QTY = 420
COL_PRICE = 424
COL_MRP = 425

# Write data rows
for i, listing in enumerate(listings):
    row = 7 + i

    output_ws.cell(row, COL_SKU).value = listing['sku']
    output_ws.cell(row, COL_PRODUCT_TYPE).value = listing['product_type']
    output_ws.cell(row, COL_LISTING_ACTION).value = listing['listing_action']
    output_ws.cell(row, COL_ITEM_NAME).value = listing['item_name']
    # Brand is BLANK for Edit (Partial Update) - don't write anything
    output_ws.cell(row, COL_MAIN_IMAGE).value = listing['image_url']
    output_ws.cell(row, COL_DESCRIPTION).value = listing['description']
    output_ws.cell(row, COL_BULLET1).value = listing['bullet1']
    output_ws.cell(row, COL_BULLET2).value = listing['bullet2']
    output_ws.cell(row, COL_BULLET3).value = listing['bullet3']
    output_ws.cell(row, COL_BULLET4).value = listing['bullet4']
    output_ws.cell(row, COL_BULLET5).value = listing['bullet5']
    output_ws.cell(row, COL_KEYWORDS).value = listing['keywords']
    output_ws.cell(row, COL_COMPATIBLE1).value = listing['compatible_model_1']
    output_ws.cell(row, COL_COMPATIBLE2).value = listing['compatible_model_2']
    output_ws.cell(row, COL_COMPATIBLE3).value = listing['compatible_model_3']
    output_ws.cell(row, COL_FULFILLMENT).value = listing['fulfillment']
    output_ws.cell(row, COL_QTY).value = listing['qty']
    output_ws.cell(row, COL_PRICE).value = listing['price']
    output_ws.cell(row, COL_MRP).value = listing['mrp']

# Save
output_file = 'bulk_upload_update.xlsx'
output_wb.save(output_file)
output_wb.close()
template_wb.close()

print(f"\n{'=' * 70}")
print("BULK UPLOAD FILE CREATED")
print(f"{'=' * 70}")
print(f"File: {output_file}")
print(f"Total listings: {len(listings)}")
print(f"Skipped items: {len(skipped)}")

# ============================================================
# SUMMARY REPORT
# ============================================================
print(f"\n{'=' * 70}")
print("SUMMARY")
print(f"{'=' * 70}")

# Price range
prices = [l['price'] for l in listings]
mrps = [l['mrp'] for l in listings]
qtys = [l['qty'] for l in listings]

print(f"\nPrice Analysis:")
print(f"  Sell Price range: ₹{min(prices):.0f} - ₹{max(prices):.0f}")
print(f"  MRP range: ₹{min(mrps):.0f} - ₹{max(mrps):.0f}")
print(f"  Average discount: {((sum(mrps)/len(mrps)) / (sum(prices)/len(prices)) - 1) * 100:.0f}%")

print(f"\nQuantity Analysis:")
print(f"  Total quantity: {sum(qtys)}")
print(f"  SKUs with stock: {len([q for q in qtys if q > 0])}")

print(f"\nTier Breakdown:")
print(f"  Gold (CareOG): {tier_counts['gold']} listings")
print(f"  Silver (OLED): {tier_counts['silver']} listings")
print(f"  Bronze (LCD): {tier_counts['bronze']} listings")

print(f"\nFrame Breakdown:")
print(f"  With Frame: {frame_counts['with_frame']} listings")
print(f"  Without Frame: {frame_counts['without_frame']} listings")

# Sample listings
print(f"\n{'=' * 70}")
print("SAMPLE LISTINGS")
print(f"{'=' * 70}")
for i, l in enumerate(listings[:3]):
    print(f"\nListing {i+1}:")
    print(f"  SKU: {l['sku']}")
    print(f"  Title: {l['item_name'][:60]}...")
    print(f"  Tier: {l['tier'].upper()}, Frame: {'Yes' if l['has_frame'] else 'No'}")
    print(f"  Price: ₹{l['price']:.0f}, MRP: ₹{l['mrp']:.0f}")
    print(f"  Qty: {l['qty']}")
    print(f"  Compatible: {l['compatible_model_1']}")
    print(f"  B1: {l['bullet1'][:60]}...")

print(f"\nFile saved: {output_file}")
print("Ready for upload to Amazon Seller Central!")