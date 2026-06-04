"""
Build a review list of SKUs with TRULY WRONG model codes in BP1.

For each flagged SKU:
  - Show current (wrong) codes
  - Look up MongoDB for the most likely CORRECT device by stricter digit-aware matching
  - Show expected correct codes side-by-side

Output: wrong_codes_review.xlsx with columns:
  SKU | title | current_bp1_codes | likely_correct_device | expected_correct_codes | confidence | notes
"""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from pymongo import MongoClient

SOURCE = 'bulk_upload_update_final_FIXED_v5.xlsx'
DEST   = 'wrong_codes_review.xlsx'
MONGO_URI = 'mongodb://dev_user:jfndACzw0ypeaNPi@ac-pn1ls1y-shard-00-00.wpdrnxc.mongodb.net:27017,ac-pn1ls1y-shard-00-01.wpdrnxc.mongodb.net:27017,ac-pn1ls1y-shard-00-02.wpdrnxc.mongodb.net:27017/repairfully?ssl=true&replicaSet=atlas-mwdddo-shard-0&authSource=admin'

# ---------- Load source ----------
wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)['Flagged']
print(f'Loaded Flagged sheet: {wb.max_row-1} rows')

# ---------- MongoDB ----------
print('Connecting to MongoDB...')
db = MongoClient(MONGO_URI, serverSelectionTimeoutMS=15000)['repairfully']
coll = db['device_model_lookup']
mongo_lookup = {}  # brandSlug -> list[(deviceName, deviceNameLower, codes)]
for d in coll.find({}, {'brandSlug':1, 'deviceName':1, 'deviceNameLower':1, 'codes':1, 'finalAssignedCode':1}):
    slug = d.get('brandSlug') or ''
    dn = d.get('deviceName') or ''
    dn_lower = d.get('deviceNameLower') or ''
    codes = d.get('codes') or []
    if not codes and d.get('finalAssignedCode'):
        codes = [d['finalAssignedCode']]
    if slug and dn_lower and codes:
        mongo_lookup.setdefault(slug, []).append((dn, dn_lower, codes))
print(f'MongoDB cache: {sum(len(v) for v in mongo_lookup.values())} devices')

BRAND_ALIASES = {'redmi':'xiaomi','mi':'xiaomi','iphone':'apple','1+':'oneplus'}

def extract_code_tokens(bp1):
    text = bp1.split('—')[0]
    tokens = re.findall(r'\b[A-Za-z0-9][A-Za-z0-9/\-]{3,}\b', text)
    return [t for t in tokens if re.search(r'[A-Za-z]', t) and re.search(r'\d', t)]

# Apply same -N strip used in regenerator
_VARIANT_SUFFIX_RE = re.compile(r'^(.+?)-(\d+)$')
def strip_variant(code):
    m = _VARIANT_SUFFIX_RE.match(code)
    if not m: return code
    base = m.group(1)
    if re.search(r'\d', base): return base
    return code

def clean_codes(codes):
    out, seen = [], set()
    for c in codes:
        c = (c or '').strip()
        if not c: continue
        s = strip_variant(c)
        if s and s.upper() not in ('CANCELLED','TBA','TBD','N/A','NONE','UNKNOWN') and s not in seen:
            out.append(s); seen.add(s)
    return out

def smart_lookup(brand, model_phrase):
    """Better match: heavily weight digit-token presence + word position."""
    if not brand or not model_phrase: return None, []
    bs = BRAND_ALIASES.get(brand.lower(), brand.lower())
    devices = mongo_lookup.get(bs, [])
    if not devices: return None, []

    target = model_phrase.lower().strip()
    target = re.sub(r'[^a-z0-9 ]+',' ',target)
    target = re.sub(r'\s+',' ',target).strip()
    # Strip brand prefix
    for p in ['samsung','apple','motorola','oneplus','oppo','vivo','realme','xiaomi','redmi','poco','mi',
              'tecno','infinix','honor','nokia','asus','lava','micromax','lenovo','sony','lg','huawei',
              'iphone','google','pixel','nothing']:
        if target.startswith(p+' '):
            target = target[len(p)+1:]; break
    if bs == 'apple' and not target.startswith('iphone'):
        target = 'iphone ' + target
    if brand.lower() in ('redmi','poco','mi') and not target.startswith(brand.lower()):
        target = brand.lower() + ' ' + target

    target_tokens = target.split()
    target_digit_tokens = set(t for t in target_tokens if any(c.isdigit() for c in t))

    # Exact match wins
    for dn, dn_lower, codes in devices:
        if dn_lower == target:
            return dn, codes

    # Score: digit-token match heavily weighted
    best = None
    for dn, dn_lower, codes in devices:
        # Skip polluted entries
        if re.search(r'\b[A-Z0-9]{4,}\b', dn) and len(dn) > 25: continue
        dn_norm = re.sub(r'[^a-z0-9 ]+',' ', dn_lower)
        dn_norm = re.sub(r'\s+',' ',dn_norm).strip()
        dn_tokens = dn_norm.split()
        dn_digit_tokens = set(t for t in dn_tokens if any(c.isdigit() for c in t))

        # Score components
        digit_overlap = len(target_digit_tokens & dn_digit_tokens)
        text_overlap  = len(set(target_tokens) & set(dn_tokens))
        # Heavy bonus if digit tokens align (e.g., "11" matches "11" not "3")
        # Heavy penalty if target has digit but matched device has DIFFERENT digit
        target_only_digits = target_digit_tokens - dn_digit_tokens
        dn_only_digits = dn_digit_tokens - target_digit_tokens
        digit_mismatch_penalty = 0
        if target_digit_tokens and not digit_overlap:
            digit_mismatch_penalty = -100  # really bad
        if target_only_digits and dn_only_digits:
            digit_mismatch_penalty -= 50

        score = digit_overlap*1000 + text_overlap*10 + digit_mismatch_penalty
        # Bonus for exact-prefix or shorter device name (more specific)
        if dn_lower == target: score += 5000
        if dn_lower.startswith(target+' ') or target.startswith(dn_lower+' '): score += 200

        if score < 0: continue
        if best is None or score > best[0]:
            best = (score, dn, codes)
    if best:
        return best[1], best[2]
    return None, []

def parse_title(title):
    m = re.search(r'compatible\s+for\s+(.+?)\s+(?:\(|with|without|fingerprint|lcd|oled|amoled|in[\s-]?cell|careog|super|display)', title, re.IGNORECASE)
    head = m.group(1).strip() if m else (title or '').replace('Compatible for ', '').strip()
    head = re.sub(r'\s*\([^)]*\)\s*$', '', head).strip()
    parts = head.split()
    return (parts[0] if parts else ''), head

# ---------- Identify TRULY WRONG SKUs ----------
print('Identifying truly wrong SKUs...')
wrong = []
for i, r in enumerate(wb.iter_rows(values_only=True)):
    if i == 0: continue
    sku = r[0]
    title = r[3] or ''
    bp1 = r[15] or ''
    issue_types = r[28] or ''

    if 'no_model_codes' in issue_types or 'multi_variant_title' in issue_types: continue
    if 'potential_device_mismatch' not in issue_types: continue

    brand, model_phrase = parse_title(title)
    current_codes = extract_code_tokens(bp1)
    if not current_codes: continue

    # Check if title digits appear in any code token
    digits_in_title = re.findall(r'\d+', model_phrase)
    if not digits_in_title: continue
    codes_concat = ' '.join(current_codes).lower()
    digit_in_codes = any(d in codes_concat for d in digits_in_title)
    if digit_in_codes: continue  # likely OK

    # This SKU has wrong codes — look up what it SHOULD be
    correct_dn, correct_codes_raw = smart_lookup(brand, model_phrase)
    correct_codes = clean_codes(correct_codes_raw) if correct_codes_raw else []

    # Confidence: if smart_lookup found a device whose name digit-aligns with title
    confidence = 'unknown'
    notes = ''
    if correct_dn:
        cdn_digits = set(re.findall(r'\d+', correct_dn))
        title_digits = set(digits_in_title)
        if title_digits & cdn_digits:
            confidence = 'high'
            notes = f'digit match: {sorted(title_digits & cdn_digits)}'
        elif not cdn_digits and not title_digits:
            confidence = 'medium'
        else:
            confidence = 'low'
            notes = f'title digits {title_digits} vs matched {cdn_digits}'
    else:
        notes = 'no good match found in MongoDB — manual lookup needed'
        confidence = 'none'

    wrong.append({
        'sku': sku,
        'title': title,
        'current_codes': ', '.join(current_codes),
        'likely_device': correct_dn or '(none found)',
        'expected_codes': ', '.join(correct_codes) if correct_codes else '(none)',
        'confidence': confidence,
        'notes': notes,
    })

print(f'Found {len(wrong)} truly wrong-code SKUs')

# ---------- Write review XLSX ----------
out_wb = Workbook()
ws = out_wb.active
ws.title = 'Wrong Codes Review'

cols = ['SKU', 'Title', 'Current BP1 Codes (WRONG)', 'Likely Correct Device (MongoDB)',
        'Expected Correct Codes', 'Confidence', 'Notes']
ws.append(cols)
for cell in ws[1]:
    cell.fill = PatternFill('solid', fgColor='F4B084')
    cell.font = Font(bold=True, color='FFFFFF')
    cell.alignment = Alignment(wrap_text=True, vertical='center')

# Color-code by confidence
CONF_COLOR = {
    'high':   'C6EFCE',   # green — clear fix available
    'medium': 'FFEB9C',   # yellow — usable
    'low':    'FFC7CE',   # pink — uncertain
    'none':   'FFC7CE',   # pink
    'unknown':'FFC7CE',
}

# Sort: by brand pattern (group Oppo Reno together, Motorola together etc)
wrong.sort(key=lambda w: (w['title'].split()[2] if len(w['title'].split()) > 2 else '', w['sku']))

for w in wrong:
    ws.append([w['sku'], w['title'], w['current_codes'], w['likely_device'],
               w['expected_codes'], w['confidence'], w['notes']])
    last = ws.max_row
    color = CONF_COLOR.get(w['confidence'], 'FFFFFF')
    for cell in ws[last]:
        cell.fill = PatternFill('solid', fgColor=color)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# Column widths
widths = {'A':18, 'B':50, 'C':40, 'D':40, 'E':45, 'F':12, 'G':45}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.row_dimensions[1].height = 30

# Summary sheet
ws2 = out_wb.create_sheet('Summary')
from collections import Counter
brand_pattern = Counter()
for w in wrong:
    parts = w['title'].split()
    brand = parts[1] if len(parts) > 1 else 'unknown'
    brand_pattern[brand] += 1
ws2.append(['Brand pattern', 'Count'])
for cell in ws2[1]:
    cell.fill = PatternFill('solid', fgColor='F4B084')
    cell.font = Font(bold=True, color='FFFFFF')
for b, c in brand_pattern.most_common():
    ws2.append([b, c])
ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 12

conf_counter = Counter(w['confidence'] for w in wrong)
ws2.append([])
ws2.append(['Confidence', 'Count'])
for cell in ws2[ws2.max_row]:
    cell.fill = PatternFill('solid', fgColor='F4B084')
    cell.font = Font(bold=True, color='FFFFFF')
for c, n in conf_counter.most_common():
    ws2.append([c, n])

out_wb.save(DEST)
print(f'\nWrote {DEST}')
print(f'  Sheet 1: Wrong Codes Review ({len(wrong)} SKUs, color-coded by confidence)')
print(f'  Sheet 2: Summary (brand patterns + confidence distribution)')
print()
print('Brand pattern of wrong codes:')
for b, c in brand_pattern.most_common():
    print(f'  {b}: {c}')
