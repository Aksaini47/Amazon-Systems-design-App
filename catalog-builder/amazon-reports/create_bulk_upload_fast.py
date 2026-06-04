"""
Amazon Bulk Listing Upload - Fast Version
Only writes the columns we need, skips template copying
"""
import openpyxl
import re

print("=" * 60)
print("BULK LISTING UPLOAD")
print("=" * 60)

# Config
CONFIG = {
    'product_type': 'PHONE_ACCESSORY',
    'listing_action': 'Edit (Partial Update)',
    'fulfillment': 'Fulfillment by Merchant (Default)',
    'description': 'Compatible mobile phone display screen replacement. Tested for quality and performance before dispatch. Professional installation recommended.',
    'warranty': 'Warranty: 7 days replacement — QC tested before dispatch. No warranty after film removal. For help visit Repairfully.com'
}

def mrp_calc(p):
    try:
        p = float(p)
        if p < 500: return round(p * 5, -2)
        elif p <= 2000: return round(p * 2, -2)
        else: return round(p * 1.43, -2)
    except: return 0

def tier(t):
    t = str(t).lower()
    if 'careog' in t: return 'gold'
    elif 'oled' in t or 'amoled' in t: return 'silver'
    return 'bronze'

def has_frame(t):
    t = str(t).lower()
    return 'with frame' in t or 'folder' in t or ' wf ' in t

def has_fp(t):
    t = str(t).lower()
    return ('fingerprint support' in t or 'oled' in t) and 'no fingerprint' not in t

def get_model(t):
    m = re.search(r'Compatible for\s+(.+?)(?:\s+(?:LCD|OLED|Display|AMOLED|CareOG)|$)', str(t), re.IGNORECASE)
    if m:
        return re.sub(r'\s*\([^)]*\)\s*', ' ', m.group(1)).strip()
    return ''

def bullets(title, t, frame, fp, model):
    fp_txt = ' — Fingerprint supported after fitting' if fp else ' — Fingerprint not supported'
    b1 = f"Compatible with {model}{fp_txt} — Verify model from Settings before ordering"

    if t == 'gold':
        b2 = "Gold Tier CareOG — Tested OG combo with premium quality. Superior reliability for repair technicians."
        b3 = "Quality Tested — 100% checked for dead pixels, color, touch response." + (" With frame for easy installation." if frame else " Screen only.")
    elif t == 'silver':
        b2 = "Silver Tier OLED — Premium display with fingerprint support. Vivid colors and better contrast." if fp else "Silver Tier OLED — Premium display quality with vivid colors."
        b3 = "OLED Quality — Original chip IC, vibrant colors, tested panel." + (" With frame for perfect fit." if frame else " Professional install recommended.")
    else:
        b2 = "Bronze Tier LCD — Budget friendly good quality display. Great value for money."
        b3 = "LCD Quality — Good color reproduction, original chip compatible, tested." + (" With frame for easy installation." if frame else " Screen only.")

    if frame:
        b4 = "With Frame Assembly — Pre-framed screen for fingerprint calibration. Ready to install." if fp else "With Frame Assembly — Screen pre-pasted on frame, ready to install."
    else:
        b4 = "Screen Only — Requires frame transfer from old screen. Adhesive included. Professional installation recommended."

    b5 = "Warranty: 7 days replacement — QC tested before dispatch. No warranty after film removal. For help visit Repairfully.com"
    return [b1, b2, b3, b4, b5]

def keywords(t, fp):
    base = "display sceen replacement mobile repair"
    if t == 'gold': tk = "careog tested combo premium quality"
    elif t == 'silver': tk = "oled amoled premium fingerprint"
    else: tk = "lcd budget friendly non fingerprint"
    return f"{base} {tk} {'fingerprint support' if fp else 'non fingerprint'}"

# Read Category Report
print("\n[1] Reading Category Report...")
wb = openpyxl.load_workbook('Category+Listings+Report_05-12-2026.xlsm', read_only=True, data_only=True)
ws = wb.active

listings = []
for row in range(7, ws.max_row + 1):
    sku = ws.cell(row, 3).value
    title = ws.cell(row, 2).value
    if not sku or not title:
        continue

    price = ws.cell(row, 424).value
    qty = ws.cell(row, 420).value
    img = ws.cell(row, 29).value

    t = tier(title)
    frame = has_frame(title)
    fp = has_fp(title)
    model = get_model(title)
    calculated_mrp = mrp_calc(price)
    b = bullets(title, t, frame, fp, model)
    kw = keywords(t, fp)

    listings.append({
        'sku': str(sku).strip(),
        'title': str(title).strip(),
        'price': price,
        'mrp': calculated_mrp,
        'qty': int(qty) if qty else 0,
        'img': img or '',
        'tier': t,
        'model': model,
        'b1': b[0], 'b2': b[1], 'b3': b[2], 'b4': b[3], 'b5': b[4],
        'kw': kw
    })

wb.close()
print(f"   {len(listings)} SKUs read")

# Create minimal output file
print("\n[2] Creating output file...")

out = openpyxl.Workbook()
out_ws = out.active
out_ws.title = 'Template'

# Headers
out_ws.cell(1, 1).value = 'SKU'
out_ws.cell(1, 2).value = 'product_type'
out_ws.cell(1, 3).value = 'listing_action'
out_ws.cell(1, 4).value = 'item_name'
out_ws.cell(1, 5).value = 'brand'
out_ws.cell(1, 6).value = 'main_image_url'
out_ws.cell(1, 7).value = 'description'
out_ws.cell(1, 8).value = 'bullet1'
out_ws.cell(1, 9).value = 'bullet2'
out_ws.cell(1, 10).value = 'bullet3'
out_ws.cell(1, 11).value = 'bullet4'
out_ws.cell(1, 12).value = 'bullet5'
out_ws.cell(1, 13).value = 'keywords'
out_ws.cell(1, 14).value = 'compatible_phone_1'
out_ws.cell(1, 15).value = 'compatible_phone_2'
out_ws.cell(1, 16).value = 'compatible_phone_3'
out_ws.cell(1, 17).value = 'fulfillment_channel'
out_ws.cell(1, 18).value = 'quantity'
out_ws.cell(1, 19).value = 'your_price'
out_ws.cell(1, 20).value = 'mrp'

# Write data
for i, l in enumerate(listings, start=2):
    out_ws.cell(i, 1).value = l['sku']
    out_ws.cell(i, 2).value = CONFIG['product_type']
    out_ws.cell(i, 3).value = CONFIG['listing_action']
    out_ws.cell(i, 4).value = l['title']
    out_ws.cell(i, 5).value = ''  # Brand BLANK for Edit
    out_ws.cell(i, 6).value = l['img']
    out_ws.cell(i, 7).value = CONFIG['description']
    out_ws.cell(i, 8).value = l['b1']
    out_ws.cell(i, 9).value = l['b2']
    out_ws.cell(i, 10).value = l['b3']
    out_ws.cell(i, 11).value = l['b4']
    out_ws.cell(i, 12).value = l['b5']
    out_ws.cell(i, 13).value = l['kw']
    out_ws.cell(i, 14).value = l['model']
    out_ws.cell(i, 15).value = ''  # Cross-fit 1
    out_ws.cell(i, 16).value = ''  # Cross-fit 2
    out_ws.cell(i, 17).value = CONFIG['fulfillment']
    out_ws.cell(i, 18).value = l['qty']
    out_ws.cell(i, 19).value = l['price']
    out_ws.cell(i, 20).value = l['mrp']

out.save('bulk_upload_update.xlsx')
out.close()

# Summary
tier_counts = {'gold': 0, 'silver': 0, 'bronze': 0}
for l in listings:
    tier_counts[l['tier']] += 1

print(f"\n{'=' * 60}")
print("DONE!")
print(f"{'=' * 60}")
print(f"File: bulk_upload_update.xlsx")
print(f"SKUs: {len(listings)}")
print(f"Gold: {tier_counts['gold']}, Silver: {tier_counts['silver']}, Bronze: {tier_counts['bronze']}")
print(f"\nSample:")
for l in listings[:3]:
    print(f"  {l['sku']}: ₹{l['price']} → MRP ₹{l['mrp']} [{l['tier'].upper()}]")
print("\nReady!")