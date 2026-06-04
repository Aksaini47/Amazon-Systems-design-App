"""
Regenerate bulk_upload_update_final.xlsx with:
  1. Column-alignment fix (insert empty other_image_url_8 cell at position 13)
  2. Title fix for MI3SG and MIRN3G — add parens around bare "Gold"
  3. Title-driven dynamic bullets per user's chat ordering:
      BP1 = Compatibility + model numbers (from gsmarena)
      BP2 = Quality tier (LCD / OLED / CareOG — language strictly follows
            bullet_point_best_practices.md eliminated-terms list)
      BP3 = Frame variant (With Frame / Without Frame)
      BP4 = Fingerprint (Fingerprint Supported / Non Fingerprint)
      BP5 = Warranty (Repairfully.com only — no GST, no WhatsApp per memory)

Language constraints (memory + best_practices.md):
  ALLOWED: LCD, OLED, CareOG, Fingerprint Supported, Non Fingerprint, With Frame,
           Without Frame, original chip IC (compound term), 7 days replacement,
           QC tested, Repairfully.com
  ELIMINATED: TFT, IPS, Incell, AMOLED, Super AMOLED, Super OLED, Super Retina XDR,
              OEM, Original (standalone), Refurbished, Premium, AAA, AAA Plus,
              Hard OLED, Soft OLED, Aftermarket, Pulled, GST, WhatsApp,
              best, #1, top-rated, refund, money back, 100% original
  TIER NAMES INTERNAL ONLY (never in bullets): Bronze, Silver, Gold

Tech mapping (title → bullet language):
  LCD              → "LCD"
  OLED             → "OLED"
  AMOLED           → "OLED"   (eliminated term mapped to allowed term)
  Super OLED       → "OLED"   (eliminated term mapped to allowed term)
  Super AMOLED     → "OLED"   (eliminated term mapped to allowed term)
  Super Retina XDR → "OLED"   (eliminated term mapped to allowed term)
  Incell           → "LCD"    (eliminated term mapped to allowed term)
  CareOG           → "CareOG"
"""
import sys, io, re, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl import Workbook

SOURCE = 'bulk_upload_update_final.xlsx'
DEST   = 'bulk_upload_update_final_FIXED_v4.xlsx'
MONGO_URI = 'mongodb://dev_user:jfndACzw0ypeaNPi@ac-pn1ls1y-shard-00-00.wpdrnxc.mongodb.net:27017,ac-pn1ls1y-shard-00-01.wpdrnxc.mongodb.net:27017,ac-pn1ls1y-shard-00-02.wpdrnxc.mongodb.net:27017/repairfully?ssl=true&replicaSet=atlas-mwdddo-shard-0&authSource=admin'
EMDASH = '—'

# -------------------- Bullet templates (language audited) --------------------
BP1_WITH_NUMS    = 'Compatible with {model} {nums} ' + EMDASH + ' Check Settings > About for exact Model Number before ordering.'
BP1_WITHOUT_NUMS = 'Compatible with {model} ' + EMDASH + ' Check Settings > About Phone for exact Model Number before ordering.'

BP2_BY_QUALITY = {
    'LCD':    '{model} LCD Display ' + EMDASH + ' replacement-grade LCD panel tested for color accuracy, dead pixels, and touch response before dispatch.',
    'OLED':   '{model} OLED Display ' + EMDASH + ' replacement-grade OLED panel tested for color accuracy, dead pixels, and touch response before dispatch.',
    'CareOG': '{model} CareOG Display ' + EMDASH + ' tested combo with original chip IC, 100% checked for dead pixels, colour calibration, and touch response before dispatch.',
}

BP3_BY_FRAME = {
    'with':    'With Frame (WF) ' + EMDASH + ' Screen pre-pasted on frame, ready to install. Adhesive pre-applied for direct fit. Colored frame rims shipped based on availability.',
    'without': 'Without Frame ' + EMDASH + ' Screen only. Requires frame transfer from your old screen. Adhesive sheet included. Professional installation recommended.',
}

BP4_BY_FP = {
    'supported':     'Fingerprint Supported ' + EMDASH + ' Under-display fingerprint sensor functional after install. Requires framing during fitting. Fingerprint takes 1-2 days to adjust after calibration.',
    'not_supported': 'Non Fingerprint ' + EMDASH + ' No under-display fingerprint sensor on this variant. Standard install with no fingerprint calibration needed.',
}

BP5_WARRANTY = 'Warranty: 7 days replacement ' + EMDASH + ' QC tested before dispatch. No warranty after film removal or installation. For help visit Repairfully.com'

# -------------------- Title fixes for bare-Gold cases --------------------
TITLE_FIXES = {
    'MI3SG':  ('Compatible for Redmi 3/3S/3X/3S Prime/3 Pro Gold LCD',  'Compatible for Redmi 3/3S/3X/3S Prime/3 Pro (Gold) LCD'),
    'MIRN3G': ('Compatible for Redmi Note 3 Gold LCD',                  'Compatible for Redmi Note 3 (Gold) LCD'),
}

# -------------------- MongoDB device lookup --------------------
print('Connecting to MongoDB device_model_lookup...')
from pymongo import MongoClient
mongo = MongoClient(MONGO_URI, serverSelectionTimeoutMS=15000)
mongo_coll = mongo['repairfully']['device_model_lookup']

def normalize_model(s):
    """Normalize for fuzzy matching: lowercase, drop non-alnum, collapse spaces."""
    s = (s or '').lower()
    s = re.sub(r'[^a-z0-9 ]+', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

# Brand aliases (SKU title brand → MongoDB brandSlug)
# Note: MongoDB has Redmi devices under brandSlug='xiaomi' with deviceName like 'Redmi 10'
BRAND_ALIASES = {
    'redmi':   'xiaomi',
    'mi':      'xiaomi',
    'iphone':  'apple',
    'oneplus': 'oneplus',
    '1+':      'oneplus',
}

# Pre-fetch all docs into local dict for fast lookup (~16.5k docs, single network round-trip)
print('Pre-fetching all device records...')
mongo_lookup = {}  # brandSlug -> list[(deviceNameLower, deviceNameNorm, codes_list)]
total = 0
for doc in mongo_coll.find({}, {'brandSlug': 1, 'deviceName': 1, 'deviceNameLower': 1, 'codes': 1, 'finalAssignedCode': 1}):
    slug = doc.get('brandSlug') or ''
    dn_lower = doc.get('deviceNameLower') or ''
    dn = doc.get('deviceName') or ''
    codes = doc.get('codes') or []
    # Fall back to finalAssignedCode if codes array is empty
    if not codes and doc.get('finalAssignedCode'):
        codes = [doc['finalAssignedCode']]
    if slug and dn_lower and codes:
        mongo_lookup.setdefault(slug, []).append((dn_lower, normalize_model(dn), codes, dn))
        total += 1
print(f'Loaded {total} device records across {len(mongo_lookup)} brands')

_VARIANT_SUFFIX_RE = re.compile(r'^(.+?)-(\d+)$')

def _strip_variant_suffix(code):
    """Strip trailing -N suffix only when prefix already contains digits.
    XT1922-1 → XT1922 (prefix XT1922 has digits, -1 is a regional variant)
    CT-888   → CT-888 (prefix CT has no digits, -888 is the actual model number)
    Mi-FX-1  → Mi-FX-1 (prefix Mi-FX has no digits, treat -1 as part of model name)
    XT2173-3 → XT2173 (prefix XT2173 has digits)
    """
    m = _VARIANT_SUFFIX_RE.match(code)
    if not m: return code
    base = m.group(1)
    if re.search(r'\d', base):
        return base
    return code

def _clean_models(models):
    """Clean gsmarena Models string:
    - Reject placeholders ('CANCELLED', 'TBA', etc.)
    - Strip trailing '-N' regional-variant suffix only when prefix has digits
      (so XT1922-1 → XT1922 but CT-888 stays CT-888).
    - Dedupe the resulting list while preserving order.
    """
    if not models: return ''
    if models.strip().upper() in ('CANCELLED', 'TBA', 'TBD', 'N/A', 'NONE', 'UNKNOWN'): return ''
    parts = [p.strip() for p in models.split(',') if p.strip()]
    if not parts: return ''
    cleaned = []
    seen = set()
    for p in parts:
        stripped = _strip_variant_suffix(p)
        if stripped and stripped not in seen:
            cleaned.append(stripped)
            seen.add(stripped)
    return ', '.join(cleaned)

def lookup_model_numbers(brand, model_str):
    """Look up clean device model codes from MongoDB device_model_lookup by brand+model.
    Returns comma-separated cleaned codes string or ''."""
    if not brand or not model_str:
        return ''
    brand_lower = brand.lower()
    brand_slug = BRAND_ALIASES.get(brand_lower, brand_lower)
    devices = mongo_lookup.get(brand_slug, [])
    if not devices:
        return ''
    # Strip brand prefix from model search term (deviceNameLower has no brand prefix)
    target = normalize_model(model_str)
    target_no_brand = target
    brand_prefixes = ['samsung', 'apple', 'motorola', 'oneplus', 'oppo', 'vivo', 'realme',
                      'xiaomi', 'redmi', 'poco', 'mi', 'tecno', 'infinix', 'honor', 'nokia',
                      'asus', 'lava', 'micromax', 'lenovo', 'sony', 'lg', 'huawei', 'iphone',
                      'google', 'pixel']
    for prefix in brand_prefixes:
        if target_no_brand.startswith(prefix + ' '):
            target_no_brand = target_no_brand[len(prefix)+1:]
            break
    # Special: Apple devices in DB have deviceName like 'iPhone 12' so re-add prefix
    if brand_slug == 'apple' and not target_no_brand.startswith('iphone'):
        target_no_brand = 'iphone ' + target_no_brand
    # Special: Redmi/Poco devices in DB still keep "Redmi" or "Poco" in deviceName
    if brand_lower in ('redmi', 'poco', 'mi') and not target_no_brand.startswith(brand_lower):
        target_no_brand = brand_lower + ' ' + target_no_brand
    if not target_no_brand:
        return ''

    # 1. Exact deviceNameLower match
    for dn_lower, dn_norm, codes, dn in devices:
        if dn_lower == target_no_brand:
            return _clean_codes_list(codes)
    # 2. Exact normalized match
    for dn_lower, dn_norm, codes, dn in devices:
        if dn_norm == target_no_brand:
            return _clean_codes_list(codes)
    # 3. Best partial match — prefer the SHORTEST deviceName that contains target,
    #    or the LONGEST target that is contained in deviceName. Penalize names with
    #    embedded codes (e.g. '10T CPH2415 CPH2413' is pollution).
    best = None  # (score, codes)
    target_words = set(target_no_brand.split())
    for dn_lower, dn_norm, codes, dn in devices:
        # Skip polluted entries that have model codes embedded in the name
        if re.search(r'\b[A-Z0-9]{4,}\b', dn) and len(dn) > 25:
            continue
        # Word-overlap score
        dn_words = set(dn_norm.split())
        if not dn_words: continue
        overlap = len(target_words & dn_words)
        if overlap == 0: continue
        # Prefer exact word match over partial
        if dn_lower == target_no_brand or dn_norm == target_no_brand:
            score = (1000, -len(dn_norm))
        elif target_no_brand.startswith(dn_lower) or dn_lower.startswith(target_no_brand):
            score = (500 + overlap, -abs(len(dn_norm) - len(target_no_brand)))
        else:
            score = (overlap, -abs(len(dn_norm) - len(target_no_brand)))
        if best is None or score > best[0]:
            best = (score, codes)
    if best:
        return _clean_codes_list(best[1])
    return ''

def _clean_codes_list(codes):
    """Apply -N variant stripping to each code in the array, dedupe, return CSV string."""
    if not codes: return ''
    cleaned = []
    seen = set()
    for c in codes:
        c = (c or '').strip()
        if not c: continue
        s = _strip_variant_suffix(c)
        if s and s.upper() not in ('CANCELLED', 'TBA', 'TBD', 'N/A', 'NONE', 'UNKNOWN') and s not in seen:
            cleaned.append(s)
            seen.add(s)
    return ', '.join(cleaned)

# -------------------- Title parsers --------------------
def detect_quality(title):
    """Return BP2-language quality tier from title text."""
    t = (title or '').lower()
    if 'careog' in t: return 'CareOG'
    # All OLED/AMOLED variants → 'OLED'
    if 'oled' in t or 'amoled' in t or 'super retina' in t: return 'OLED'
    # Default and Incell → 'LCD'
    return 'LCD'

def detect_frame(title, sku):
    t = (title or '').lower()
    if 'with frame' in t or 'with-frame' in t: return 'with'
    if 'without frame' in t or 'no frame' in t: return 'without'
    s = (sku or '').upper()
    if re.search(r'WF', s): return 'with'
    if re.search(r'NF', s): return 'without'
    return 'without'  # default for generic replacement screens

def detect_fingerprint(title, old_bullet1):
    t = (title or '').lower()
    if 'fingerprint not' in t or 'no fingerprint' in t: return 'not_supported'
    if 'fingerprint support' in t or 'fingerprint enabled' in t: return 'supported'
    b1 = (old_bullet1 or '').lower()
    if 'fingerprint not' in b1: return 'not_supported'
    if 'fingerprint support' in b1 or 'fingerprint enabled' in b1: return 'supported'
    return 'not_supported'  # safer default for unknown

# Brand+Model extraction from title
TITLE_HEAD_RE = re.compile(
    r'compatible\s+for\s+(.+?)\s+'
    r'(?:\(|with\s+frame|without\s+frame|fingerprint|lcd|oled|amoled|in[\s-]?cell|careog|super\s+(?:retina|oled|amoled)|display)',
    re.IGNORECASE
)
def extract_brand_model(title):
    """Return (brand, model_phrase) where model_phrase includes brand for display."""
    t = (title or '').strip()
    m = TITLE_HEAD_RE.search(t)
    if m:
        head = m.group(1).strip()
    elif t.lower().startswith('compatible for '):
        head = t[15:].strip()
    else:
        head = t
    # Strip trailing parens like "(Gold)" or "(5G)" — keep them out of the model string
    head = re.sub(r'\s*\([^)]*\)\s*$', '', head).strip()
    # First word (or two for known multi-word brands) = brand
    parts = head.split()
    if not parts:
        return '', head
    multi_word_brands = {('one', 'plus'): 'OnePlus'}
    brand = parts[0]
    return brand, head

# -------------------- Main --------------------
print(f'Loading source: {SOURCE}')
src_wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
src_ws = src_wb['Template']
src_rows = list(src_ws.iter_rows(values_only=True))
header = list(src_rows[0])
data_rows = src_rows[1:]
print(f'Loaded {len(data_rows)} data rows. Source header: {len(header)} cols.')

# Empirical column map for SOURCE rows (data shifted -1 from col 13+)
SRC = {
    'SKU': 0, 'product_type': 1, 'listing_action': 2, 'item_name': 3, 'brand': 4,
    'main_image_url': 5,
    'other_image_url_1': 6, 'other_image_url_2': 7, 'other_image_url_3': 8,
    'other_image_url_4': 9, 'other_image_url_5': 10, 'other_image_url_6': 11,
    'other_image_url_7': 12,
    'description': 13, 'bullet1': 14, 'bullet2': 15, 'bullet3': 16, 'bullet4': 17, 'bullet5': 18,
    'keywords': 19,
    'compatible_phone_1': 20, 'compatible_phone_2': 21, 'compatible_phone_3': 22,
    'fulfillment_channel': 23, 'quantity': 24, 'your_price': 25, 'mrp': 26,
}

# Canonical 28-col DEST layout matching the header
DEST_HEADER = [
    'SKU', 'product_type', 'listing_action', 'item_name', 'brand',
    'main_image_url',
    'other_image_url_1', 'other_image_url_2', 'other_image_url_3',
    'other_image_url_4', 'other_image_url_5', 'other_image_url_6',
    'other_image_url_7', 'other_image_url_8',
    'description', 'bullet1', 'bullet2', 'bullet3', 'bullet4', 'bullet5',
    'keywords',
    'compatible_phone_1', 'compatible_phone_2', 'compatible_phone_3',
    'fulfillment_channel', 'quantity', 'your_price', 'mrp',
]
assert len(DEST_HEADER) == 28

dst_wb = Workbook()
dst_ws = dst_wb.active
dst_ws.title = 'Template'
dst_ws.append(DEST_HEADER)

# Audit counters
stats = {
    'quality_LCD': 0, 'quality_OLED': 0, 'quality_CareOG': 0,
    'frame_with': 0, 'frame_without': 0,
    'fp_supported': 0, 'fp_not_supported': 0,
    'title_fixed': 0, 'rows_processed': 0,
    'over_255': 0, 'over_1000_bytes': 0,
    'model_numbers_found': 0, 'model_numbers_missing': 0,
}

length_issues = []
example_rows = []  # capture first 3 of each variant for review

for r in data_rows:
    sku = r[SRC['SKU']] or ''
    title = r[SRC['item_name']] or ''
    brand_col = r[SRC['brand']]            # blank on Edit Partial Update — preserve
    desc  = r[SRC['description']] or ''
    keywords = r[SRC['keywords']] or ''

    # Title fix for bare-Gold SKUs
    if sku in TITLE_FIXES:
        old, new = TITLE_FIXES[sku]
        if old in title:
            title = title.replace(old, new)
            stats['title_fixed'] += 1

    # Parse title attributes
    old_b1 = r[SRC['bullet1']] or ''
    quality = detect_quality(title)
    frame   = detect_frame(title, sku)
    fp      = detect_fingerprint(title, old_b1)
    brand_str, model_phrase = extract_brand_model(title)

    # Lookup model numbers from gsmarena (best-effort)
    nums = lookup_model_numbers(brand_str, model_phrase)
    if nums:
        stats['model_numbers_found'] += 1
    else:
        stats['model_numbers_missing'] += 1

    # Generate bullets
    if nums:
        bp1 = BP1_WITH_NUMS.format(model=model_phrase, nums=nums)
    else:
        bp1 = BP1_WITHOUT_NUMS.format(model=model_phrase)
    bp2 = BP2_BY_QUALITY[quality].format(model=model_phrase)
    bp3 = BP3_BY_FRAME[frame]
    bp4 = BP4_BY_FP[fp]
    bp5 = BP5_WARRANTY

    # Length checks
    bullets = [bp1, bp2, bp3, bp4, bp5]
    for i, b in enumerate(bullets, 1):
        if len(b) > 255:
            length_issues.append((sku, f'bp{i}', len(b)))
            stats['over_255'] += 1
            # Trim BP1 if too long (model_numbers can be huge)
            if i == 1:
                bp1 = BP1_WITHOUT_NUMS.format(model=model_phrase)
                bullets[0] = bp1
    bullets_total_bytes = len(' '.join(bullets).encode('utf-8'))
    if bullets_total_bytes > 1000:
        length_issues.append((sku, 'bullets_total_bytes', bullets_total_bytes))
        stats['over_1000_bytes'] += 1

    stats[f'quality_{quality}'] = stats.get(f'quality_{quality}', 0) + 1
    stats[f'frame_{frame}'] = stats.get(f'frame_{frame}', 0) + 1
    stats[f'fp_{fp}'] = stats.get(f'fp_{fp}', 0) + 1

    # Capture example rows
    key = f'{quality}/{frame}/{fp}'
    if sum(1 for ex in example_rows if ex[0] == key) < 1:
        example_rows.append((key, sku, title, bp1, bp2, bp3, bp4, bp5))

    # Build canonical 28-col destination row
    out = [None] * 28
    out[0]  = sku
    out[1]  = r[SRC['product_type']]
    out[2]  = r[SRC['listing_action']]
    out[3]  = title
    out[4]  = brand_col   # blank — preserved per Edit Partial Update standard
    out[5]  = r[SRC['main_image_url']]
    out[6]  = r[SRC['other_image_url_1']]
    out[7]  = r[SRC['other_image_url_2']]
    out[8]  = r[SRC['other_image_url_3']]
    out[9]  = r[SRC['other_image_url_4']]
    out[10] = r[SRC['other_image_url_5']]
    out[11] = r[SRC['other_image_url_6']]
    out[12] = r[SRC['other_image_url_7']]
    out[13] = None   # other_image_url_8 — empty cell to align with header
    out[14] = desc
    out[15] = bullets[0]
    out[16] = bullets[1]
    out[17] = bullets[2]
    out[18] = bullets[3]
    out[19] = bullets[4]
    out[20] = keywords
    out[21] = r[SRC['compatible_phone_1']]
    out[22] = r[SRC['compatible_phone_2']]
    out[23] = r[SRC['compatible_phone_3']]
    out[24] = r[SRC['fulfillment_channel']]
    out[25] = r[SRC['quantity']]
    out[26] = r[SRC['your_price']]
    out[27] = r[SRC['mrp']]
    dst_ws.append(out)
    stats['rows_processed'] += 1

dst_wb.save(DEST)

print(f'\n=== STATS ===')
for k, v in stats.items():
    print(f'  {k}: {v}')

if length_issues:
    print(f'\n=== LENGTH ISSUES ({len(length_issues)}) ===')
    for sku, kind, n in length_issues[:20]:
        print(f'  {sku}: {kind} = {n}')

print(f'\n=== EXAMPLE BULLETS (one per variant) ===')
for key, sku, title, b1, b2, b3, b4, b5 in example_rows[:12]:
    print(f'\n[{key}] SKU={sku}')
    print(f'  title: {title}')
    print(f'  BP1:   {b1}')
    print(f'  BP2:   {b2}')
    print(f'  BP3:   {b3}')
    print(f'  BP4:   {b4}')
    print(f'  BP5:   {b5}')

print(f'\nWrote {DEST}')
