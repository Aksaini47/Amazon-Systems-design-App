"""
v6 regenerator: MongoDB-driven, Indian-variant aware, smart digit-matching.

Flow per SKU:
  1. Parse brand + model from title (e.g. "Oppo Reno 4")
  2. MongoDB lookup with DIGIT-WEIGHTED matching (so "Reno 11" matches Reno 11 entry,
     not generic Reno)
  3. Extract:
       - indiaCode      = MongoDB.indiaModelCode OR finalAssignedCode (single India variant)
       - allCodes       = MongoDB.codes array (all global variants, cleaned)
  4. TITLE = "Compatible for [Brand] [Model] ([indiaCode]) [other parens] [screen] Display..."
       (single India code in parens, right after model name, before other attribute parens)
  5. BP1   = "Compatible with [Brand] [Model] ([code1, code2, code3...]) — Check Settings > About
              for exact Model Number before ordering."
       (all codes comma-separated in single parens)
  6. BP2/3/4/5 = unchanged from v3 templates (LCD/OLED/CareOG quality, frame, fp, warranty)

Constraints:
  - TITLE ≤ 200 chars
  - BP1 ≤ 255 chars (truncate code list if needed; keep India code first + as many as fit)
  - Multi-variant titles (with "/"): use first variant's codes
  - Column-alignment fix preserved (canonical 28-col output)
"""
import sys, io, re
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from openpyxl import Workbook
from pymongo import MongoClient

SOURCE = 'bulk_upload_update_final.xlsx'
DEST   = 'bulk_upload_update_final_FIXED_v6.xlsx'
MONGO_URI = 'mongodb://dev_user:jfndACzw0ypeaNPi@ac-pn1ls1y-shard-00-00.wpdrnxc.mongodb.net:27017,ac-pn1ls1y-shard-00-01.wpdrnxc.mongodb.net:27017,ac-pn1ls1y-shard-00-02.wpdrnxc.mongodb.net:27017/repairfully?ssl=true&replicaSet=atlas-mwdddo-shard-0&authSource=admin'
EMDASH = '—'

# ---------- Bullet templates (unchanged from v3) ----------
BP1_TEMPLATE = 'Compatible with {model} ({codes}) ' + EMDASH + ' Check Settings > About for exact Model Number before ordering.'
BP1_NO_CODES = 'Compatible with {model} ' + EMDASH + ' Check Settings > About Phone for exact Model Number before ordering.'

BP2_BY_QUALITY = {
    'LCD':    '{model} LCD Display ' + EMDASH + ' replacement-grade LCD panel tested for color accuracy, dead pixels, and touch response before dispatch.',
    'OLED':   '{model} OLED Display ' + EMDASH + ' replacement-grade OLED panel tested for color accuracy, dead pixels, and touch response before dispatch.',
    'CareOG': '{model} CareOG Display ' + EMDASH + ' tested combo with original chip IC, 100% checked for dead pixels, colour calibration, and touch response before dispatch.',
}
BP3_BY_FRAME = {
    'with':    'With Frame (WF) ' + EMDASH + ' Screen pre-pasted on frame, ready to install. Adhesive pre-applied for direct fit. Colored frame rims shipped based on availability.',
    'without': 'Without Frame ' + EMDASH + ' Screen only. Requires frame transfer from your old screen. Adhesive sheet included. Professional installation recommended.',
}
BP4_BY_FP = {
    'supported':     'Fingerprint Supported ' + EMDASH + ' Under-display fingerprint sensor functional after install. Requires framing during fitting. Fingerprint takes 1-2 days to adjust after calibration.',
    'not_supported': 'Non Fingerprint ' + EMDASH + ' No under-display fingerprint sensor on this variant. Standard install with no fingerprint calibration needed.',
}
BP5_WARRANTY = 'Warranty: 7 days replacement ' + EMDASH + ' QC tested before dispatch. No warranty after film removal or installation. For help visit Repairfully.com'

# Title fixes
TITLE_FIXES = {
    'MI3SG':  ('Compatible for Redmi 3/3S/3X/3S Prime/3 Pro Gold LCD',  'Compatible for Redmi 3/3S/3X/3S Prime/3 Pro (Gold) LCD'),
    'MIRN3G': ('Compatible for Redmi Note 3 Gold LCD',                  'Compatible for Redmi Note 3 (Gold) LCD'),
}

# Brand aliases
BRAND_ALIASES = {'redmi':'xiaomi','mi':'xiaomi','iphone':'apple','1+':'oneplus'}

# ---------- Load MongoDB ----------
print('Loading MongoDB device lookup...')
db = MongoClient(MONGO_URI, serverSelectionTimeoutMS=20000)['repairfully']
coll = db['device_model_lookup']

mongo_lookup = {}  # brandSlug -> list[(deviceName, deviceNameLower, codes, indiaCode, finalCode)]
for d in coll.find({}, {'brandSlug':1, 'deviceName':1, 'deviceNameLower':1,
                        'codes':1, 'finalAssignedCode':1, 'indiaModelCode':1}):
    slug = d.get('brandSlug') or ''
    dn = d.get('deviceName') or ''
    dn_lower = d.get('deviceNameLower') or ''
    codes = d.get('codes') or []
    final = d.get('finalAssignedCode') or ''
    india = d.get('indiaModelCode') or ''
    if slug and dn_lower:
        mongo_lookup.setdefault(slug, []).append({
            'dn': dn, 'dn_lower': dn_lower, 'codes': codes,
            'india': india, 'final': final
        })
print(f'  Loaded {sum(len(v) for v in mongo_lookup.values())} devices across {len(mongo_lookup)} brands')

# ---------- Code cleaner (strip -N variants only when prefix has digits) ----------
_VARIANT_SUFFIX_RE = re.compile(r'^(.+?)-(\d+)$')
def _strip_variant(c):
    m = _VARIANT_SUFFIX_RE.match(c)
    if not m: return c
    base = m.group(1)
    return base if re.search(r'\d', base) else c

_QUALIFIER_TOKENS = {'DS','DSN','DUAL','SIM','DDS','TDS','/DS','BANDS','GSM','LTE','5G','4G','3G'}

def _is_qualifier(c):
    """True if token is a qualifier/marker, not a real model code.
    Qualifiers: <=3 char pure letters (DS, DSN), or known marker words."""
    if not c: return True
    up = c.upper().strip('/').strip('-')
    if up in _QUALIFIER_TOKENS: return True
    # 2-3 char pure-letter tokens are typically qualifiers (DS, FN, DSN, etc.)
    if len(c) <= 3 and c.isalpha(): return True
    return False

def _clean_code(c):
    """Clean a single code: strip -N variant suffix, return stripped string (or '' if invalid)."""
    if not c: return ''
    c = c.strip()
    if not c: return ''
    if c.upper() in ('CANCELLED','TBA','TBD','N/A','NONE','UNKNOWN'): return ''
    if _is_qualifier(c): return ''
    return _strip_variant(c)

def _clean_codes(codes):
    """Clean list of codes: reject placeholders+qualifiers, strip -N variants, dedupe."""
    if not codes: return []
    out, seen = [], set()
    for c in codes:
        s = _clean_code(c)
        if s and s not in seen:
            out.append(s); seen.add(s)
    return out

# ---------- Smart digit-aware MongoDB lookup ----------
def _norm(s):
    s = (s or '').lower()
    s = re.sub(r'[^a-z0-9 ]+', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

BRAND_PREFIXES = ['samsung','apple','motorola','oneplus','oppo','vivo','realme','xiaomi','redmi','poco','mi',
                  'tecno','infinix','honor','nokia','asus','lava','micromax','lenovo','sony','lg','huawei',
                  'iphone','google','pixel','nothing']

def smart_lookup(brand, model_phrase):
    """Returns dict with dn, india_code, all_codes — or None if no good match.
    DIGIT-WEIGHTED: 'Reno 11' must match Reno 11 entry (digit '11'), not generic 'Reno'."""
    if not brand or not model_phrase: return None
    bs = BRAND_ALIASES.get(brand.lower(), brand.lower())
    devices = mongo_lookup.get(bs, [])
    if not devices: return None

    target = _norm(model_phrase)
    for p in BRAND_PREFIXES:
        if target.startswith(p + ' '):
            target = target[len(p)+1:]; break
    # Re-add brand prefix for matches that use it
    if bs == 'apple' and not target.startswith('iphone'):
        target = 'iphone ' + target
    if brand.lower() in ('redmi','poco','mi') and not target.startswith(brand.lower()):
        target = brand.lower() + ' ' + target

    target_tokens = set(target.split())
    target_digits = set(t for t in target.split() if any(c.isdigit() for c in t))

    # 1. Exact match
    for d in devices:
        if d['dn_lower'] == target:
            india = d['india'] or d['final'] or ''
            return {'dn': d['dn'], 'india_code': india, 'all_codes': _clean_codes(d['codes'])}

    # 2. Score-based: heavy digit-token weighting, exact-digit-match required if title has digits
    best = None
    for d in devices:
        dn_lower = d['dn_lower']
        # Skip polluted/very long entries
        if len(d['dn']) > 30 and re.search(r'\b[A-Z0-9]{4,}\b', d['dn']): continue
        dn_norm = _norm(dn_lower)
        dn_tokens = set(dn_norm.split())
        dn_digits = set(t for t in dn_norm.split() if any(c.isdigit() for c in t))

        digit_overlap = len(target_digits & dn_digits)
        token_overlap = len(target_tokens & dn_tokens)

        # HARD penalty: if target has digit tokens and NONE align with matched, reject
        if target_digits and not (target_digits & dn_digits):
            continue
        # HARD penalty: if matched device has different digits than target (e.g. target "11" vs matched "3")
        target_only = target_digits - dn_digits
        dn_only = dn_digits - target_digits
        if target_only and dn_only:
            # Allow only if target's digit is contained within matched (e.g. "10" in "10 Pro")
            # i.e. digit subset is OK, digit mismatch is bad
            if not (target_digits <= dn_digits or dn_digits <= target_digits):
                continue

        score = digit_overlap * 10000 + token_overlap * 100
        # Exact word-set match
        if dn_tokens == target_tokens: score += 50000
        # Target is a prefix of dn (dn is more specific): light penalty (we want the right one)
        if dn_lower.startswith(target + ' '): score += 500
        # Closer length = better match
        score -= abs(len(dn_lower) - len(target)) * 2

        if best is None or score > best[0]:
            best = (score, d)

    if best:
        d = best[1]
        india = d['india'] or d['final'] or ''
        return {'dn': d['dn'], 'india_code': india, 'all_codes': _clean_codes(d['codes'])}
    return None

# ---------- Title parsers ----------
def detect_quality(title):
    t = (title or '').lower()
    if 'careog' in t: return 'CareOG'
    if 'oled' in t or 'amoled' in t or 'super retina' in t: return 'OLED'
    return 'LCD'

def detect_frame(title, sku):
    t = (title or '').lower()
    if 'with frame' in t or 'with-frame' in t: return 'with'
    if 'without frame' in t or 'no frame' in t: return 'without'
    s = (sku or '').upper()
    if re.search(r'WF', s): return 'with'
    if re.search(r'NF', s): return 'without'
    return 'without'

def detect_fingerprint(title, old_b1):
    t = (title or '').lower()
    if 'fingerprint not' in t or 'no fingerprint' in t: return 'not_supported'
    if 'fingerprint support' in t or 'fingerprint enabled' in t: return 'supported'
    b1 = (old_b1 or '').lower()
    if 'fingerprint not' in b1: return 'not_supported'
    if 'fingerprint support' in b1 or 'fingerprint enabled' in b1: return 'supported'
    return 'not_supported'

TITLE_HEAD_RE = re.compile(
    r'compatible\s+for\s+(.+?)\s+'
    r'(?:\(|with\s+frame|without\s+frame|fingerprint|lcd|oled|amoled|in[\s-]?cell|careog|super\s+(?:retina|oled|amoled)|display)',
    re.IGNORECASE
)
def extract_brand_model(title):
    """Return (brand, full_model_phrase_incl_brand, primary_model_for_lookup)."""
    t = (title or '').strip()
    m = TITLE_HEAD_RE.search(t)
    head = m.group(1).strip() if m else (t.replace('Compatible for ', '').strip())
    head = re.sub(r'\s*\([^)]*\)\s*$', '', head).strip()
    parts = head.split()
    brand = parts[0] if parts else ''
    # For multi-variant titles ("S1/S1 Pro", "A50/A50s", "Redmi 3/3S/3X/3S Prime/3 Pro"),
    # use the first variant for lookup (keep full display).
    if '/' in head:
        primary = head.split('/')[0].strip()
        # If the after-slash part contains a space (e.g. "S1 Pro"), take only the bit before
        primary_for_lookup = primary
    else:
        primary_for_lookup = head
    return brand, head, primary_for_lookup

def inject_code_into_title(title, india_code):
    """Insert ' (INDIA_CODE)' right after the model name, before any other parens
    or screen-type keyword. Returns modified title."""
    if not india_code: return title
    # Find the position right after the model name. The model phrase is followed by either:
    # (a) an existing paren, e.g. "Galaxy A30s (Fingerprint Support)"
    # (b) a screen-type keyword, e.g. "OnePlus 5 OLED"
    # (c) "Display" or similar
    pat = re.compile(
        r'(compatible\s+for\s+.+?)'
        r'(\s+(?:\(|with\s+frame|without\s+frame|fingerprint|lcd|oled|amoled|in[\s-]?cell|careog|super\s+(?:retina|oled|amoled)|display)\b)',
        re.IGNORECASE
    )
    m = pat.search(title)
    if not m:
        # Fallback: append at end
        return f'{title} ({india_code})'
    head = m.group(1)
    rest = m.group(2)
    insertion = f' ({india_code})'
    new_title = head + insertion + rest + title[m.end():]
    return new_title

# ---------- Load source ----------
print(f'Loading source: {SOURCE}')
src_wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
src_ws = src_wb['Template']
src_rows = list(src_ws.iter_rows(values_only=True))
header = list(src_rows[0])
data_rows = src_rows[1:]
print(f'  {len(data_rows)} SKUs loaded.')

# Empirical SRC mapping (data shifted -1 from col 13+)
SRC = {
    'SKU':0,'product_type':1,'listing_action':2,'item_name':3,'brand':4,
    'main_image_url':5,'other_image_url_1':6,'other_image_url_2':7,'other_image_url_3':8,
    'other_image_url_4':9,'other_image_url_5':10,'other_image_url_6':11,'other_image_url_7':12,
    'description':13,'bullet1':14,'bullet2':15,'bullet3':16,'bullet4':17,'bullet5':18,
    'keywords':19,'compatible_phone_1':20,'compatible_phone_2':21,'compatible_phone_3':22,
    'fulfillment_channel':23,'quantity':24,'your_price':25,'mrp':26,
}

# Canonical DEST header
DEST_HEADER = [
    'SKU','product_type','listing_action','item_name','brand',
    'main_image_url',
    'other_image_url_1','other_image_url_2','other_image_url_3',
    'other_image_url_4','other_image_url_5','other_image_url_6',
    'other_image_url_7','other_image_url_8',
    'description','bullet1','bullet2','bullet3','bullet4','bullet5',
    'keywords','compatible_phone_1','compatible_phone_2','compatible_phone_3',
    'fulfillment_channel','quantity','your_price','mrp',
]

dst_wb = Workbook()
dst_ws = dst_wb.active
dst_ws.title = 'Template'
dst_ws.append(DEST_HEADER)

# ---------- Length-aware code fitter for BP1 ----------
def fit_codes_to_bp1(model, codes, max_chars=255):
    """Build BP1 with as many codes as fit in 255 chars. India code goes first."""
    if not codes:
        return BP1_NO_CODES.format(model=model), 0
    # Try all codes
    full_bp = BP1_TEMPLATE.format(model=model, codes=', '.join(codes))
    if len(full_bp) <= max_chars:
        return full_bp, len(codes)
    # Trim from end until fits
    n = len(codes)
    while n > 1:
        n -= 1
        trimmed = codes[:n]
        bp = BP1_TEMPLATE.format(model=model, codes=', '.join(trimmed))
        if len(bp) <= max_chars:
            return bp, n
    # Just one code
    return BP1_TEMPLATE.format(model=model, codes=codes[0]), 1

# ---------- Process ----------
stats = {'rows':0, 'title_fixed':0, 'codes_found':0, 'codes_missing':0,
         'india_in_title':0, 'multi_variant':0,
         'over_255':0, 'over_1000':0, 'title_over_200':0,
         'quality_LCD':0, 'quality_OLED':0, 'quality_CareOG':0,
         'frame_with':0, 'frame_without':0, 'fp_supported':0, 'fp_not_supported':0}

samples = []   # one per variant
for r in data_rows:
    sku = r[SRC['SKU']] or ''
    title = r[SRC['item_name']] or ''
    brand_col = r[SRC['brand']]
    desc = r[SRC['description']] or ''
    keywords = r[SRC['keywords']] or ''

    if sku in TITLE_FIXES:
        old, new = TITLE_FIXES[sku]
        if old in title:
            title = title.replace(old, new)
            stats['title_fixed'] += 1

    old_b1 = r[SRC['bullet1']] or ''
    quality = detect_quality(title)
    frame   = detect_frame(title, sku)
    fp      = detect_fingerprint(title, old_b1)
    brand, full_model, lookup_model = extract_brand_model(title)
    if full_model != lookup_model:
        stats['multi_variant'] += 1

    # MongoDB lookup
    match = smart_lookup(brand, lookup_model)
    india_code_raw = (match.get('india_code') if match else '') or ''
    india_code = _clean_code(india_code_raw)  # apply -N strip + qualifier filter to India code too
    all_codes = match.get('all_codes') if match else []

    if all_codes:
        stats['codes_found'] += 1
    else:
        stats['codes_missing'] += 1

    # Reorder codes: india_code first, then others (dedup)
    if india_code and india_code in all_codes:
        codes_for_bp1 = [india_code] + [c for c in all_codes if c != india_code]
    elif india_code:
        codes_for_bp1 = [india_code] + all_codes
    else:
        codes_for_bp1 = all_codes

    # Inject India code into title
    new_title = title
    if india_code:
        new_title = inject_code_into_title(title, india_code)
        stats['india_in_title'] += 1
        if len(new_title) > 200:
            # Revert if title would exceed limit
            new_title = title
            stats['india_in_title'] -= 1
    # Check title length
    if len(new_title) > 200:
        stats['title_over_200'] += 1

    # Build bullets
    bp1, codes_used = fit_codes_to_bp1(full_model, codes_for_bp1)
    bp2 = BP2_BY_QUALITY[quality].format(model=full_model)
    bp3 = BP3_BY_FRAME[frame]
    bp4 = BP4_BY_FP[fp]
    bp5 = BP5_WARRANTY

    bullets = [bp1, bp2, bp3, bp4, bp5]
    for i, b in enumerate(bullets, 1):
        if len(b) > 255:
            stats['over_255'] += 1
    total_bytes = len(' '.join(bullets).encode('utf-8'))
    if total_bytes > 1000:
        stats['over_1000'] += 1

    stats['rows'] += 1
    stats[f'quality_{quality}'] = stats.get(f'quality_{quality}', 0) + 1
    stats[f'frame_{frame}'] = stats.get(f'frame_{frame}', 0) + 1
    stats[f'fp_{fp}'] = stats.get(f'fp_{fp}', 0) + 1

    key = f'{quality}/{frame}/{fp}'
    if sum(1 for s in samples if s[0] == key) < 1:
        samples.append((key, sku, new_title, bp1, bp2, bp3, bp4, bp5, match))

    # Build dest row
    out = [None]*28
    out[0]  = sku
    out[1]  = r[SRC['product_type']]
    out[2]  = r[SRC['listing_action']]
    out[3]  = new_title
    out[4]  = brand_col
    out[5]  = r[SRC['main_image_url']]
    out[6]  = r[SRC['other_image_url_1']]
    out[7]  = r[SRC['other_image_url_2']]
    out[8]  = r[SRC['other_image_url_3']]
    out[9]  = r[SRC['other_image_url_4']]
    out[10] = r[SRC['other_image_url_5']]
    out[11] = r[SRC['other_image_url_6']]
    out[12] = r[SRC['other_image_url_7']]
    out[13] = None     # other_image_url_8 — empty for alignment
    out[14] = desc
    out[15] = bullets[0]
    out[16] = bullets[1]
    out[17] = bullets[2]
    out[18] = bullets[3]
    out[19] = bullets[4]
    out[20] = keywords
    out[21] = r[SRC['compatible_phone_1']]
    out[22] = r[SRC['compatible_phone_2']]
    out[23] = r[SRC['compatible_phone_3']]
    out[24] = r[SRC['fulfillment_channel']]
    out[25] = r[SRC['quantity']]
    out[26] = r[SRC['your_price']]
    out[27] = r[SRC['mrp']]
    dst_ws.append(out)

dst_wb.save(DEST)

print(f'\n=== STATS ===')
for k, v in stats.items():
    print(f'  {k}: {v}')

print(f'\n=== SAMPLE BULLETS (one per variant) ===')
for key, sku, t, b1, b2, b3, b4, b5, m in samples[:12]:
    print(f'\n[{key}] {sku}')
    print(f'  TITLE: {t}')
    if m: print(f'  matched device: "{m.get("dn")}" | india: {m.get("india_code")} | codes: {m.get("all_codes")[:6]}')
    print(f'  BP1:   {b1}')
    print(f'  BP2:   {b2}')
    print(f'  BP3:   {b3}')
    print(f'  BP4:   {b4}')
    print(f'  BP5:   {b5}')

print(f'\nWrote {DEST}')
